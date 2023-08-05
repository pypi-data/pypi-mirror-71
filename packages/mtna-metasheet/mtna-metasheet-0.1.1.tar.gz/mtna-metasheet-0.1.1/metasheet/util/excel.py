"""Metasheet Excel parser"""
import re
import openpyxl

from . import repository
from .classes import Resource

def readExcel(workspace, filename):
    """Parses an Excel file"""
    wb = openpyxl.load_workbook(filename, data_only=True)
    parseWorkbook(workspace,wb)

def parseWorkbook(workspace,wb):
    """Parses all sheets matching entities in an Excel workbook"""
    # PARSE ALL THE SHEETS IN THE WORKBOOK
    print("LOADING WORKBOOK...")
    for type in repository.getResourceTypes().keys():
        regex = repository.getResourceTypeSheetRegex(type)
        for name in wb.sheetnames:
            if re.match(regex, name, re.IGNORECASE):
                sheet = wb[name]
                parseSheet(workspace,type, sheet)
    print("PARSING COMPLETED")

def parseSheet(workspace,resourceType, sheet):
    """Parses an Excel sheet"""
    print()
    print(sheet, "-->", resourceType)

    # Setup column-->property mappings
    maps = []  # this array hold a map for each column in the sheet, or None if not found
    for row in sheet.iter_rows(max_row=1): # first row
        for index, cell in enumerate(row): # first row
            propertyMap = None
            print(cell)
            if cell.value:
                # check entity specific maps
                if repository.getPropertyMaps(resourceType):
                    for map in repository.getPropertyMaps(resourceType):
                        match = repository.matchMap(map, cell.value)
                        if match:
                            propertyMap = {"name":match,"map":map}
                            break
                # layout can also use all the variable maps
                if propertyMap is None and resourceType == "layout":
                    print("Layout using Variable maps")
                    for map in repository.getPropertyMaps("variable"):
                        match = repository.matchMap(map, cell.value)
                        if match:
                            propertyMap = {"name":match,"map":map}
                            break
                # check global maps
                if propertyMap is None:
                    for map in repository.getPropertyMaps():
                        match = repository.matchMap(map, cell.value)
                        if match:
                            propertyMap = {"name":match,"map":map}
                            break
            print("Column", index, ":", cell.value, "-->", propertyMap)
            maps.append(propertyMap)

    # read resource metadata
    count = 0
    for row in sheet.iter_rows(min_row=2): # skip first row
        if str(row[0].value).startswith("#"):
            # comment row (first cell starts with #)
            continue

        # create a new resource
        source = str(sheet) + " row " + str(row[0].row)
        resource = Resource(resourceType, source)
        # add properties
        for index, cell in enumerate(row):
            if maps[index] is not None and cell.value:
                # drop decimals zeros (e.g. 1.0 --> 1)
                # to address integers stored as float issue
                if isinstance(cell.value, float) and cell.value.is_integer:
                    cell.value = int(cell.value)
                # add the property
                resource.addProperty(maps[index].get("name"),cell.value,maps[index].get("map"))
        # register the resource
        if resource.getProperties():
            workspace.addResource(resource)
            count += 1
            if len(resource.getProperties())  <= 2:
                print(len(resource.getProperties()))
                resource.dump()
        else:
            # empty resource
            pass
    
    print(count)

 