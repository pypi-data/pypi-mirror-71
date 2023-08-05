""" utility.py

Utility functions. """

from __future__ import annotations
from typing import Optional, Dict, List, Union
from os import path
from datetime import datetime
from openpyxl import load_workbook
from json import dump, dumps, load
from annif_client import AnnifClient

from .jskos import _Concept, _ConceptScheme, _LanguageMap


class _Utility:
    """ A collection of utility functions. """

    @classmethod
    def load_file(cls, filename: str, language: str = "und") -> Optional[_ConceptScheme]:
        """ Load a file.

        :param filename: the name of the file including its complete path
        :param language: the language of the words given as RFC 3066 language tag, defaults to "und"
        """

        # stop if file does not exist
        if path.exists(filename) is False:
            print(f"ERROR: File {filename} does not exist!")
            return None

        # choose method depending on file type:
        if filename.endswith(".xlsx"):
            workbook = load_workbook(filename)
            return cls.xlsx2scheme(workbook, language)
        elif filename.endswith(".json"):
            # TODO: add JSON support
            pass
        else:
            pass

    @classmethod
    def xlsx2scheme(cls, workbook, language: str = "und") -> _ConceptScheme:
        """ Transform a XLSX workbook into a JSKOS concept scheme.

        The XLSX workbook's data structure MUST be as follows: one column with one row per word.

        :param workbook: the XLSX workbook
        :param language: the language of the words given as RFC 3066 language tag, defaults to "und"
        """

        words = []
        for worksheet in workbook:
            for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
                if row[0] is None:
                    continue
                else:
                    words.append(row[0])

        return cls.words2scheme(words=words, language=language)

    @classmethod
    def words2scheme(cls,
                     words: List[str],
                     uri: str = None,
                     name: str = None,
                     language: str = "und") -> _ConceptScheme:
        """ Transform a list of words into a JSKOS concept scheme.

        :param words: the input words
        :param uri: the URI of the concept scheme, defaults to None
        :param name: the name of the concept scheme, defaults to None
        :param language: the language of the concept scheme given as RFC 3066 language tag, defaults to "und"
        """

        # make the concept scheme:
        if name is None:
            name = f"session-{datetime.now().toordinal()}"
        if uri is None:
            uri = f"bartocsuggest:scheme/{name}?language={language}"
        scheme = _ConceptScheme(uri=uri)  # to include name: pref_label=_LanguageMap({language.lower(): name}

        # add the concepts:
        counter = 1
        for word in words:
            concept = cls.word2concept(word=word, scheme_uri=scheme.uri, language=language, notation=str(counter))
            scheme.concepts.append(concept)
            counter =+ 1

        return scheme

    @classmethod
    def word2concept(cls, word: str, scheme_uri: str, language: str = "und", notation: str = None) -> _Concept:
        """ Transform a word into a JSKOS concept.

        :param word: the input word
        :param scheme_uri: the URI of the concept scheme
        :param language: the language of the word given as RFC 3066 language tag, defaults to "und"
        :param notation: the notation of the word, defaults to None
        """

        uri = cls.word2uri(word, language)
        concept = _Concept(uri=uri,
                           pref_label=_LanguageMap({language.lower(): word}),
                           in_scheme=set([scheme_uri]))
        if notation is not None:
            concept.notation = [notation]

        return concept

    @classmethod
    def word2uri(cls, word: str, language: str = "und") -> str:
        """ Transform a word into a URI.

        :param word: the input word
        :param language: the language of the word given as RFC 3066 language tag, defaults to "und"
        """

        replacements = {" ": "-", "/": "-"}
        for key, value in replacements.items():
            word = word.replace(key, value)

        return f"bartocsuggest:concept/{word}?language={language}"

    @classmethod
    def annif2jskos(cls, annif_suggestion: List[dict], annif_project_id: str) -> _ConceptScheme:
        """ Transform an Annif suggestion into a JSKOS concept scheme.

        :param annif_suggestion: the output of calling AnnifClient.suggest
        :param annif_project_id: Annif API project identifier
        """

        # get project details:
        annif = AnnifClient()
        project = annif.get_project(annif_project_id)
        language = project.get("language")
        name = project.get("name")

        # setup concept scheme based on project details:
        scheme = _ConceptScheme(pref_label=_LanguageMap({language: name}))

        # make concept from result and add to concept scheme:
        for result in annif_suggestion:
            label = result.get("label")
            uri = result.get("uri")
            concept = _Concept(uri=uri, pref_label=_LanguageMap({language: label}))
            scheme.concepts.append(concept)

        return scheme

    @classmethod
    def save_json(cls, dictionary: Dict, folder: str, filename: str = None):
        """ Save a dictionary as JSON file.

        :param dictionary: the dictionary to be saved
        :param folder: the folder to write the JSON file in (MUST use complete folder path)
        :param filename: the name of the file, defaults to None
        """

        if filename is None:
            filename = str(datetime.now()).split(".")[0].replace(":", "-")

        full_filename = folder + f"{filename}.json"
        with open(full_filename, "w") as file:
            dump(dictionary, file)

    @classmethod
    def load_json(cls, folder: str, filename: str) -> Dict:
        """ Load a JSON object as dictionary from a file.

        :param folder: the path to the folder
        :param filename: the name of the file
        """

        filename = folder + f"{filename}.json"

        with open(filename) as file:
            dictionary = load(file)

            return dictionary

    @classmethod
    def print_json(cls, dictionary: dict, indent: int = 2) -> None:
        """ Print a dictionary as JSON to the console.

        :param dictionary: the dictionary to be printed
        :param indent: indentation for pretty printing, defaults to 2
        """

        print(dumps(dictionary, indent=indent))
