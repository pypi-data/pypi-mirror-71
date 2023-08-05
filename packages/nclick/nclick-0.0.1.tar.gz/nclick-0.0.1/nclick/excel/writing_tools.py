from copy import copy

import openpyxl as px
from openpyxl.utils.dataframe import dataframe_to_rows

from nclick.excel import cell_styles as cs


def get_location_symbol(row_idx,
                        col_idx):
    """
    セルの位置情報を表す文字列を返す（例: "B12"）

    Parameters
    ----------
    row_idx : integer
        行方向の添字
    col_idx : integer
        列方向の添字

    Returns
    ----------
    location_symbol : string
        位置情報を表す文字列（例: "B12"）

    Notes
    ----------
    Excelの添字は１から始まることに注意
  """
    location_symbol = f'{px.utils.get_column_letter(col_idx)}{row_idx}'
    return location_symbol

def set_value(ws,
              value,
              idx_row,
              idx_col):
    """
    セルに値を代入

    Parameters
    ----------
    ws : openpyxl worksheet
        値を代入するExcel Sheet
    value :
        代入する値
    idx_row : integer
        代入するセルの行番号
    idx_col : integer
        代入するセルの列番号
    """
    cell_loc = get_location_symbol(idx_row, idx_col)
    ws[cell_loc].value = value

def rotate_text(ws,
                rotate,
                idx_row,
                idx_col):
    """
    文字列を回転

    Parameters
    ----------
    ws : openpyxl worksheet
        styleを適用するExcel Sheet
    rotate : integer
        回転する角度
    idx_row : integer
        styleを適用するセルの行番号
    idx_col : integer
        styleを適用するセルの列番号
    """
    cell_loc = get_location_symbol(idx_row, idx_col)

    tmp_alignment = copy(ws[cell_loc].alignment)
    tmp_alignment.text_rotation=rotate
    ws[cell_loc].alignment = tmp_alignment

def set_style(ws,
              style,
              idx_row,
              idx_col):
    """
    セルに値を代入

    Parameters
    ----------
    ws : openpyxl worksheet
        styleを適用するExcel Sheet
    idx_row : integer
        styleを適用するセルの行番号
    idx_col : integer
        styleを適用するセルの列番号
    """
    cell_loc = get_location_symbol(idx_row, idx_col)
    ws[cell_loc].style = style

def set_conditinal_format(ws,
                          conditinal_format,
                          start_row,
                          end_row,
                          start_col,
                          end_col):
    """
    選択範囲（矩形のみ）のセルにstyleを適用

    Parameters
    ----------
    ws : openpyxl worksheet
        関数を適用するExcel Sheet
    conditinal_format : openpyxl conditinal format
        条件付きフォーマット
    start_row : integer
        選択範囲のセルにおける行番号の下限値
    end_row : integer
        選択範囲のセルにおける行番号の上限値
    start_col : integer
        選択範囲のセルにおける列番号の下限値
    end_col : integer
        選択範囲のセルにおける列番号の上限値
    """
    start_loc = get_location_symbol(start_row, start_col)
    end_loc = get_location_symbol(end_row, end_col)
    ws.conditional_formatting.add(f'{start_loc}:{end_loc}', conditinal_format)

def fill_cell(ws,
              start_row,
              end_row,
              start_col,
              end_col,
              named_style):
    """
    選択範囲（矩形のみ）のセルにstyleを適用

    Parameters
    ----------
    ws : openpyxl worksheet
        関数を適用するExcel Sheet
    start_row : integer
        選択範囲のセルにおける行番号の下限値
    end_row : integer
        選択範囲のセルにおける行番号の上限値
    start_col : integer
        選択範囲のセルにおける列番号の下限値
    end_col : integer
        選択範囲のセルにおける列番号の上限値
    named_style : style
        適用するスタイル
    """
    style_name = named_style.name
    for i, idx_row in enumerate(range(start_row, end_row)):
        for j, idx_col in enumerate(range(start_col, end_col)):
            if i == 0 and j == 0:
                set_style(ws, named_style, idx_row, idx_col)
            else:
                set_style(ws, style_name, idx_row, idx_col)

def set_number_format(ws,
                      start_row,
                      end_row,
                      start_col,
                      end_col,
                      number_format):
    """
    選択範囲（矩形のみ）のセルに数値フォーマットを適用

    Parameters
    ----------
    ws : openpyxl worksheet
        関数を適用するExcel Sheet
    start_row : integer
        選択範囲のセルにおける行番号の下限値
    end_row : integer
        選択範囲のセルにおける行番号の上限値
    start_col : integer
        選択範囲のセルにおける列番号の下限値
    end_col : integer
        選択範囲のセルにおける列番号の上限値
    number_format : string
        適用するスタイル
    """
    for idx_row in range(start_row, end_row):
        for idx_col in range(start_col, end_col):
            cell_loc = get_location_symbol(idx_row, idx_col)
            ws[cell_loc].number_format = number_format

def dataframe_to_sheet(df,
                       ws,
                       start_row=1,
                       start_col=1,
                       header=True,
                       index_cols_num=0,
                       header_style=cs.style_00,
                       index_style=cs.style_01,
                       cell_style=cs.style_02,
                       number_format={}):
    """
    データフレームを書き出す

    Parameters
    ----------
    df : pandas DataFrame
        Excelに書き出すデータフレーム
    ws : openpyxl worksheet
        データフレームを書き出すExcel Sheet
    start_row : integer
        左上に配置されるセルの行番号
    start_col : integer
        左上に配置されるセルの列番号
    header : bool
        headerを書き出すか否か
    index_cols_num : integer
        indexとして扱うカラムの数
    header_style :
        headerのstyle
    index_style :
        indexのstyle
    cell_style :
        cellのstyle
    """
    n_rows, n_cols = df.shape

    # 値の書き出し
    rows = dataframe_to_rows(df, index=False, header=header)
    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            set_value(ws, value, start_row+r_idx, start_col+c_idx)

    # header styleを適用
    if header and header_style is not None:
        fill_cell(ws,
                  start_row,
                  start_row+1,
                  start_col,
                  start_col+n_cols,
                  header_style)

    # index styleを適用
    if index_cols_num >0 and index_style is not None:
        fill_cell(ws,
                  start_row+1,
                  start_row+1+n_rows,
                  start_col,
                  start_col+index_cols_num,
                  index_style)

    # cell styleを適用
    if cell_style is not None:
        fill_cell(ws,
                  start_row+1,
                  start_row+1+n_rows,
                  start_col+index_cols_num,
                  start_col+n_cols,
                  cell_style)

    # カラムごとに数値フォーマットを適用
    if type(number_format) == dict and len(number_format) != 0:
        cols = df.columns.tolist()
        for col, num_format in number_format.items():
            col_idx = cols.index(col)
            set_number_format(ws,
                              start_row+1,
                              start_row+1+n_rows,
                              start_col+col_idx,
                              start_col+col_idx+1,
                              num_format)

    # 全セルに数値フォーマットを適用
    if type(number_format) == str:
        set_number_format(ws,
                          start_row+1,
                          start_row+1+n_rows,
                          start_col,
                          start_col+n_cols,
                          number_format)

def set_column_width(ws,
                     col_idx,
                     width):
    """
    指定したカラムの幅を調整

    Parameters
    ----------
    ws : openpyxl worksheet
        自動調節するExcel Sheet
    col_idx : integer
        列のインデックス
    width : float
        カラムの幅
    """
    col_symbol = px.utils.get_column_letter(col_idx)
    ws.column_dimensions[col_symbol].width = width

def set_row_height(ws,
                  row_idx,
                  height):
    """
    指定した行の幅を調整

    Parameters
    ----------
    ws : openpyxl worksheet
        自動調節するExcel Sheet
    row_idx : integer
        行のインデックス
    height : float
        行の幅
    """
    ws.row_dimensions[row_idx].height = height

def autoresize_columns_width(ws,
                             ratio=1.5):
    """
    セルの幅を自動調整

    Parameters
    ----------
    ws : openpyxl worksheet
        自動調節するExcel Sheet
    ratio : float
        幅の拡大率
    """
    for col in ws.columns:
        unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, col))
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * ratio
        ws.column_dimensions[unmerged_cells[0].column_letter].width = adjusted_width

def paste_image(ws,
                image_path,
                start_row,
                start_col):
    """
    画像ファイルを貼り付ける

    Parameters
    ----------
    ws : openpyxl worksheet
        画像を貼り付けるExcel Sheet
    start_row : integer
        画像の左上にあたるセルの行番号
    start_col : integer
        画像の左上にあたるセルの列番号
    """
    img = px.drawing.image.Image(image_path)
    img_loc = get_location_symbol(start_row, start_col)
    ws.add_image(img, img_loc)

def freeze_header_and_index(ws,
                            idx_row,
                            idx_col):
    """
    headerとindexを固定

    Parameters
    ----------
    ws : openpyxl worksheet
        固定するheader, indexを持つExcel Sheet
    idx_row : integer
        固定する行の行番号
    idx_col : integer
        固定する列の列番号
    """
    cell_loc = get_location_symbol(idx_row, idx_col)
    ws.freeze_panes = cell_loc

def merge_cells(ws,
                start_row,
                end_row,
                start_col,
                end_col):
    """
    セルをマージ

    Parameters
    ----------
    ws : openpyxl worksheet
        マージ対象のセルを持つExcel Sheet
    start_row : integer
        結合後のセルにおける行番号の下限値
    end_row : integer
        結合後のセルにおける行番号の上限値
    start_col : integer
        結合後のセルにおける列番号の下限値
    end_col : integer
        結合後のセルにおける列番号の上限値
    """
    ws.merge_cells(start_row=start_row,
                   start_column=start_col,
                   end_row=end_row,
                   end_column=end_col)