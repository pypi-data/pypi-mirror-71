from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule


# 薄緑（主にheader用)
style_00 = NamedStyle(name="style_00",
                      font=Font(bold=True, size=10.5),
                      fill=PatternFill(patternType='solid', start_color='E6EDDD', end_color='E6EDDD'),
                      border=Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin',color='000000'),
                                    top=Side(border_style='thin',color='000000'),
                                    bottom=Side(border_style='thin', color='000000')),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))

# グレー（主にindex用)
style_01 = NamedStyle(name="style_01",
                      font=Font(bold=True, size=10.5),
                      fill=PatternFill(patternType='solid', start_color='F2F2F2', end_color='F2F2F2'),
                      border=Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin',color='000000'),
                                    top=Side(border_style='thin',color='000000'),
                                    bottom=Side(border_style='thin', color='000000')),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))

# 塗りつぶしなし（主にcell用)
style_02 = NamedStyle(name="style_02",
                      font=Font(bold=False, size=10.5),
                      border=Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin',color='000000'),
                                    top=Side(border_style='thin',color='000000'),
                                    bottom=Side(border_style='thin', color='000000')),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))


# 淡赤（主に強調したいcell用)
style_03 = NamedStyle(name="style_03",
                      border=Border(left=Side(border_style=None),
                                    right=Side(border_style=None),
                                    top=Side(border_style=None),
                                    bottom=Side(border_style=None)),
                      font=Font(bold=True, size=10.5),
                      fill=PatternFill(patternType='solid', start_color='EEDDDC', end_color='EEDDDC'),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))

# 淡青（主に強調したいcell用)
style_04 = NamedStyle(name="style_04",
                      border=Border(left=Side(border_style=None),
                                    right=Side(border_style=None),
                                    top=Side(border_style=None),
                                    bottom=Side(border_style=None)),
                      font=Font(bold=True, size=10.5),
                      fill=PatternFill(patternType='solid', start_color='DEE6F0', end_color='DEE6F0'),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))

# 淡緑（主に強調したいcell用)
style_05 = NamedStyle(name="style_05",
                      border=Border(left=Side(border_style=None),
                                    right=Side(border_style=None),
                                    top=Side(border_style=None),
                                    bottom=Side(border_style=None)),
                      font=Font(bold=True, size=10.5),
                      fill=PatternFill(patternType='solid', start_color='E6EDDD', end_color='E6EDDD'),
                      alignment=Alignment(horizontal='center',
                                          vertical = 'center'))

# グラデーション（緑）
colorscale_00 = ColorScaleRule(start_type='min',
                               start_color='e5f5f9',
                               end_type='max',
                               end_color='2ca25f')

# グラデーション（赤）
colorscale_01 = ColorScaleRule(start_type='min',
                               start_color='fff7ec',
                               end_type='max',
                               end_color='990000')

# グラデーション（青->赤）
colorscale_02 = ColorScaleRule(start_type='num',
                               start_value=-1,
                               start_color='0571b0',
                               mid_type='num',
                               mid_value=0,
                               mid_color='f7f7f7',
                               end_type='num',
                               end_value=1,
                               end_color='ca0020')