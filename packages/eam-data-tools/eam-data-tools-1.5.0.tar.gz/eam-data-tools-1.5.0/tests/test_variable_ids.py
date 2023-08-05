import unittest
import openpyxl
from table_data_reader import OpenpyxlTableHandler
from deepdiff import DeepDiff


def get_diff_then_make_files_equal(a, b):
    wb_a = openpyxl.load_workbook(a, data_only=True)
    wb_b = openpyxl.load_workbook(b, data_only=True)
    diff = DeepDiff(wb_b, wb_a)
    sheet_names = wb_a.sheetnames
    for sheet_name in sheet_names:
        sheet_a = wb_a[sheet_name]
        sheet_b = wb_b[sheet_name]
        for row in range(len(list(sheet_b.iter_rows()))):
            for col in range(len(list(sheet_b.iter_cols()))):
                c = sheet_a.cell(row=row + 1, column=col + 1)
                c.value = sheet_b.cell(row=row + 1, column=col + 1).value
    wb_a.save(a)
    return diff


class TestVariableIDs(unittest.TestCase):

    def test_existing_ids(self):
        handler = OpenpyxlTableHandler()
        handler.load_definitions("params", filename="tests/data/existing_ids.xlsx", id_flag=True)
        diff = get_diff_then_make_files_equal("tests/data/existing_ids.xlsx", "tests/data/existing_ids_copy.xlsx")
        assert handler.id_map == {'power_latop': {'n/a': 0, "S1": 1},
                                  'energy_intensity_network': {'n/a': 2}}
        assert diff == {}

    def test_some_existing_ids(self):
        handler = OpenpyxlTableHandler()
        handler.load_definitions("params", filename="tests/data/some_existing_ids.xlsx", id_flag=True)
        diff = get_diff_then_make_files_equal("tests/data/some_existing_ids.xlsx",
                                              "tests/data/some_existing_ids_copy.xlsx")
        assert handler.id_map == {'power_latop': {'n/a': 0}, 'time_laptop': {'n/a': 3},
                                  'energy_intensity_network': {'n/a': 2}}
        print(type(diff))
        assert (str(
            diff) == "{'type_changes': {'root[0][2][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 3}, 'root[0][2][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}}}"
                or str(
                diff == "{'type_changes': {'root[0][2][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}, 'root[0][2][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 3}}}"))

    def test_no_existing_ids(self):
        handler = OpenpyxlTableHandler()
        handler.load_definitions("params", filename="tests/data/no_existing_ids.xlsx", id_flag=True)
        diff = get_diff_then_make_files_equal("tests/data/no_existing_ids.xlsx",
                                              "tests/data/no_existing_ids_copy.xlsx")
        assert handler.id_map == {'power_latop': {'n/a': 0 , 'S1' :1}}
        print(diff)
        assert (str(
            diff) == "{'type_changes': {'root[0][1][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 0}, 'root[0][1][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}, 'root[0][2][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 1}, 'root[0][2][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}}}"
                or str(diff) == "{'type_changes': {'root[0][1][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}, 'root[0][1][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 0}, 'root[0][2][18]._style': {'old_type': <class 'NoneType'>, 'new_type': <class 'openpyxl.styles.cell_style.StyleArray'>, 'old_value': None, 'new_value': StyleArray('i', [0, 0, 0, 0, 0, 0, 0, 0, 0])}, 'root[0][2][18]._value': {'old_type': <class 'NoneType'>, 'new_type': <class 'int'>, 'old_value': None, 'new_value': 1}}}")

    def test_duplicate_ids(self):
        with self.assertRaises(Exception) as context:
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params", filename="tests/data/duplicate_ids.xlsx", id_flag=True)
        self.assertTrue("Duplicate ID variable " in str(context.exception))

    def test_no_id_flag(self):
        handler = OpenpyxlTableHandler()
        handler.load_definitions("params", filename="tests/data/no_existing_ids.xlsx")
        diff = get_diff_then_make_files_equal("tests/data/no_existing_ids.xlsx",
                                              "tests/data/no_existing_ids_copy.xlsx")
        assert diff == {}
        assert handler.id_map == {}
