__version__ = '1.0.0'

import csv
import datetime
import importlib

from abc import abstractmethod
from collections import defaultdict
from typing import Dict, List, Set

import numpy as np
import pandas as pd
from dateutil import relativedelta as rdelta

import logging
from functools import partial

import calendar

from openpyxl import Workbook
from scipy.interpolate import interp1d
import json
from typing import Callable

__author__ = 'schien'

# import pkg_resources  # part of setuptools
# version = pkg_resources.require("excel-modelling-helper")[0].version

param_name_map_v1 = {'variable': 'name', 'scenario': 'source_scenarios_string', 'module': 'module_name',
                     'distribution': 'distribution_name', 'param 1': 'param_a', 'param 2': 'param_b',
                     'param 3': 'param_c',
                     'unit': '', 'CAGR': 'cagr', 'ref date': 'ref_date', 'label': '', 'tags': '', 'comment': '',
                     'source': ''}

param_name_map_v2 = {'CAGR': 'cagr',
                     'comment': '',
                     'label': '',
                     'mean growth': 'growth_factor',
                     'param': '',
                     'ref date': 'ref_date',
                     'ref value': '',
                     'scenario': 'source_scenarios_string',
                     'source': '',
                     'tags': '',
                     'type': '',
                     'unit': '',
                     'variability growth': 'ef_growth_factor',
                     'initial_value_proportional_variation': '',
                     'variable': 'name'}

param_name_maps = {1: param_name_map_v1, 2: param_name_map_v2}

# logger.basicConfig(level=logger.DEBUG)
logger = logging.getLogger(__name__)


class DistributionFunctionGenerator(object):
    module: str
    distribution: str
    param_a: str
    param_b: str
    param_c: str

    def __init__(self, module_name=None, distribution_name=None, param_a: float = None,
                 param_b: float = None, param_c: float = None, size=None, **kwargs):
        """
        Instantiate a new object.

        :param module_name:
        :param distribution_name:
        :param param_a:
        :param param_b:
        :param param_c:
        :param size:
        :param kwargs: can contain key "sample_mean_value" with bool value
        """
        self.kwargs = kwargs
        self.size = size
        self.module_name = module_name
        self.distribution_name = distribution_name
        self.sample_mean_value = kwargs.get('sample_mean_value', False)
        # prepare function arguments
        if distribution_name == 'choice':
            if type(param_a) == str:
                tokens = param_a.split(',')
                params = [float(token.strip()) for token in tokens]
                self.random_function_params = [np.array(params, dtype=np.float)]
            else:
                self.random_function_params = [np.array([i for i in [param_a, param_b, param_c] if i], dtype=np.float)]

            logger.debug(f'setting function params for choice distribution {self.random_function_params}')
        else:
            self.random_function_params = [i for i in [param_a, param_b, param_c] if i not in [None, ""]]

    def get_mean(self, distribution_function):
        """Get the mean value for a distribution.
        If the distribution function is [normal, uniform,choice,triangular] the analytic value is being calculted.
        Else, the distribution is instantiated and then the mean is being calculated.

        :param distribution_function:
        :return: the mean as a scalar
        """
        name = self.distribution_name
        params = self.random_function_params
        if name == 'normal':
            return params[0]
        if name == 'uniform':
            return (params[0] + params[1]) / 2.
        if name == 'choice':
            return params[0].mean()
        if name == 'triangular':
            return (params[0] + params[1] + params[2]) / 3.
        return distribution_function().mean()

    def generate_values(self, *args, **kwargs):
        """
        Generate a sample of values by sampling from a distribution. The size of the sample can be overriden with the 'size' kwarg.

        If `self.sample_mean_value == True` the sample will contain "size" times the mean value.

        :param args:
        :param kwargs:
        :return: sample as vector of given size
        """
        sample_size = kwargs.get('size', self.size)

        f = self.instantiate_distribution_function(self.module_name, self.distribution_name)
        distribution_function = partial(f, *self.random_function_params, size=sample_size)

        if self.sample_mean_value:
            sample = np.full(sample_size, self.get_mean(distribution_function))
        else:
            sample = distribution_function()

        return sample

    @staticmethod
    def instantiate_distribution_function(module_name, distribution_name):
        module = importlib.import_module(module_name)
        func = getattr(module, distribution_name)
        return func


class Parameter(object):
    """
    A single parameter
    """
    version: int

    name: str
    unit: str
    comment: str
    source: str
    scenario: str

    processes: Dict[str, List]

    "optional comma-separated list of tags"
    tags: str

    def __init__(self, name, tags=None, source_scenarios_string: str = None, unit: str = None,
                 comment: str = None, source: str = None, version=None,
                 **kwargs):
        self.version = version
        # The source definition of scenarios. A comma-separated list
        self.source = source
        self.comment = comment

        self.unit = unit
        self.source_scenarios_string = source_scenarios_string
        self.tags = tags
        self.name = name

        self.scenario = None
        self.cache = None

        # track the usages of this parameter per process as a list of process-specific variable names that are backed by this parameter
        self.processes = defaultdict(list)

        self.kwargs = kwargs

    def __call__(self, settings=None, *args, **kwargs):
        """
        Samples from a parameter. Values are cached and returns the same value every time called.

        @todo confusing interface that accepts 'settings' and kwargs  at the same time.
        worse- 'use_time_series' must be present in the settings dict

        :param args:
        :param kwargs:
        :return:
        """
        if self.cache is None:
            kwargs['name'] = self.name
            kwargs['unit'] = self.unit
            kwargs['tags'] = self.tags
            kwargs['scenario'] = self.scenario

            if not settings:
                settings = {}

            common_args = {'size': settings.get('sample_size', 1),
                           'sample_mean_value': settings.get('sample_mean_value', False),
                           'with_pint_units': settings.get('with_pint_units', False)
                           }
            common_args.update(**self.kwargs)

            if settings.get('use_time_series', False):
                if self.version == 2:
                    generator = GrowthTimeSeriesGenerator(**common_args, times=settings['times'])
                else:
                    generator = ConstantUncertaintyExponentialGrowthTimeSeriesGenerator(**common_args,
                                                                                        times=settings['times'])
            else:
                generator = DistributionFunctionGenerator(**common_args)

            self.cache = generator.generate_values(*args, **kwargs)
        return self.cache

    def add_usage(self, process_name, variable_name):
        # add the name of a variable of a process model that is backed by this parameter
        self.processes[process_name].append(variable_name)


class GrowthTimeSeriesGenerator(DistributionFunctionGenerator):
    ref_date: str
    # of the mean values
    # the type of growth ['exp']
    # growth_function_type: str
    # of the error function
    variance: str
    # error function growth rate
    ef_growth_factor: str

    def __init__(self, times=None, size=None, index_names=None, ref_date=None, with_pint_units=False, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.ref_date = ref_date if ref_date else None
        self.with_pint_units = with_pint_units
        if self.with_pint_units:
            import pint
        self.times = times
        self.size = size
        iterables = [times, range(0, size)]
        self._multi_index = pd.MultiIndex.from_product(iterables, names=index_names)
        assert type(times.freq) == pd.tseries.offsets.MonthBegin, 'Time index must have monthly frequency'

    def generate_values(self, *args, **kwargs):
        """
        Instantiate a random variable and apply annual growth factors.

        :return:
        """
        assert 'ref value' in self.kwargs

        # 1. Generate $\mu$
        start_date = self.times[0].to_pydatetime()
        end_date = self.times[-1].to_pydatetime()
        ref_date = self.ref_date
        if not ref_date:
            raise Exception(f"Ref date not set for variable {kwargs['name']}")

        mu = self.generate_mu(end_date, ref_date, start_date)

        # 3. Generate $\sigma$
        # Prepare array with growth values $\sigma$
        if self.sample_mean_value:
            sigma = np.zeros((len(self.times), self.size))
        else:

            if self.kwargs['type'] == 'interp':

                def get_date(record):
                    return datetime.datetime.strptime(record[0], "%Y-%m-%d")

                ref_value_ = sorted(json.loads(self.kwargs['ref value'].strip()).items(), key=get_date)
                intial_value = ref_value_[0][1]
            else:
                intial_value = float(self.kwargs['ref value'])

            variability_ = intial_value * self.kwargs['initial_value_proportional_variation']
            logger.debug(f'sampling random distribution with parameters -{variability_}, 0, {variability_}')
            sigma = np.random.triangular(-1 * variability_, 0, variability_, (len(self.times), self.size))
        # logger.debug(ref_date.strftime("%b %d %Y"))

        # 4. Prepare growth array for $\alpha_{sigma}$
        alpha_sigma = growth_coefficients(start_date,
                                          end_date,
                                          ref_date,
                                          self.kwargs['ef_growth_factor'], 1)

        # 5. Prepare DataFrame
        iterables = [self.times, range(self.size)]
        index_names = ['time', 'samples']
        _multi_index = pd.MultiIndex.from_product(iterables, names=index_names)

        # logger.debug(start_date)
        # logger.debug(end_date)
        from dateutil import relativedelta
        r = relativedelta.relativedelta(end_date, start_date)
        months = r.years * 12 + r.months + 1
        name = kwargs['name']
        # Apply growth to $\sigma$ and add $\sigma$ to $\mu$
        # logger.debug(sigma.size)
        # logger.debug(alpha_sigma.shape)
        # logger.debug(months)
        if self.with_pint_units:
            unit_ = kwargs["unit"]
            if not unit_:
                unit_ = 'dimensionless'
            dtype = f'pint[{unit_}]'
        else:
            dtype = 'float64'

        series = pd.Series(((sigma * alpha_sigma) + mu.reshape(months, 1)).ravel(), index=_multi_index,
                           dtype=dtype)

        # test if df has sub-zero values
        df_sigma__dropna = series.where(series < 0).dropna()

        if self.with_pint_units:
            _values = df_sigma__dropna.pint.m
        else:
            _values = df_sigma__dropna

        if not _values.empty:
            logger.warning(f"Negative values for parameter {name} from {df_sigma__dropna.index[0][0]}")

        return series

    def generate_mu(self, end_date, ref_date, start_date):

        if self.kwargs['type'] == 'exp':
            mu_bar = np.full(len(self.times), float(self.kwargs['ref value']))
            # 2. Apply Growth to Mean Values $\alpha_{mu}$
            alpha_mu = growth_coefficients(start_date,
                                           end_date,
                                           ref_date,
                                           self.kwargs['growth_factor'], 1)
            mu = mu_bar * alpha_mu.ravel()
            mu = mu.reshape(len(self.times), 1)
            return mu
        if self.kwargs['type'] == 'interp':
            def toTimestamp(d):
                return calendar.timegm(d.timetuple())

            def interpolate(growth_config: Dict[str, float], date_range, kind='linear'):
                arr1 = np.array([toTimestamp(datetime.datetime.strptime(date_val, '%Y-%m-%d')) for date_val in
                                 growth_config.keys()])
                arr2 = np.array([val for val in growth_config.values()])

                f = interp1d(arr1, arr2, kind=kind, fill_value='extrapolate')
                return f([toTimestamp(date_val) for date_val in date_range])

            ref_value_ = json.loads(self.kwargs['ref value'].strip())
            return interpolate(ref_value_, self.times, self.kwargs['param'])


class ConstantUncertaintyExponentialGrowthTimeSeriesGenerator(DistributionFunctionGenerator):
    cagr: str
    ref_date: str

    def __init__(self, cagr=None, times=None, size=None, index_names=None, ref_date=None, with_pint_units=False, *args,
                 **kwargs):
        super().__init__(*args, **kwargs)
        self.cagr = cagr if cagr else 0

        self.ref_date = ref_date if ref_date else None
        self.with_pint_units = with_pint_units
        if self.with_pint_units:
            import pint
        self.times = times
        self.size = size
        iterables = [times, range(0, size)]
        self._multi_index = pd.MultiIndex.from_product(iterables, names=index_names)
        assert type(times.freq) == pd.tseries.offsets.MonthBegin, 'Time index must have monthly frequency'

    def generate_values(self, *args, **kwargs):
        """
        Instantiate a random variable and apply annual growth factors.

        :return:
        """
        values = super().generate_values(*args, **kwargs, size=(len(self.times) * self.size,))
        alpha = self.cagr

        # @todo - fill to cover the entire time: define rules for filling first
        ref_date = self.ref_date if self.ref_date else self.times[0].to_pydatetime()
        # assert ref_date >= self.times[0].to_pydatetime(), 'Ref date must be within variable time span.'
        # assert ref_date <= self.times[-1].to_pydatetime(), 'Ref date must be within variable time span.'

        start_date = self.times[0].to_pydatetime()
        end_date = self.times[-1].to_pydatetime()

        a = growth_coefficients(start_date, end_date, ref_date, alpha, self.size)

        x = a.ravel()
        values = np.multiply(values, x)

        # df = pd.DataFrame(values)
        # df.columns = [kwargs['name']]
        # df.set_index(self._multi_index, inplace=True)
        # # @todo this is a hack to return a series with index as I don't know how to set an index and rename a series
        # data_series = df.iloc[:, 0]
        # data_series._metadata = kwargs
        # data_series.index.rename(['time', 'samples'], inplace=True)
        #
        if self.with_pint_units:
            if not kwargs["unit"]:
                dtype = 'pint[dimensionless]'
            else:
                dtype = f'pint[{kwargs["unit"]}]'
        else:
            dtype = 'float64'

        series = pd.Series(values, index=self._multi_index, dtype=dtype)
        return series


def growth_coefficients(start_date, end_date, ref_date, alpha, samples):
    """
    Build a matrix of growth factors according to the CAGR formula  y'=y0 (1+a)^(t'-t0).

    a growth rate alpha
    t0 start date
    t' end date
    y' output
    y0 start value

    """

    start_offset = 0
    if ref_date < start_date:
        offset_delta = rdelta.relativedelta(start_date, ref_date)
        start_offset = offset_delta.months + 12 * offset_delta.years
        start_date = ref_date

    end_offset = 0
    if ref_date > end_date:
        offset_delta = rdelta.relativedelta(ref_date, end_date)
        end_offset = offset_delta.months + 12 * offset_delta.years
        end_date = ref_date

    delta_ar = rdelta.relativedelta(ref_date, start_date)
    ar = delta_ar.months + 12 * delta_ar.years
    delta_br = rdelta.relativedelta(end_date, ref_date)
    br = delta_br.months + 12 * delta_br.years

    # we place the ref point on the lower interval (delta_ar + 1) but let it start from 0
    # in turn we let the upper interval start from 1
    g = np.fromfunction(lambda i, j: np.power(1 - alpha, np.abs(i) / 12), (ar + 1, samples), dtype=float)
    h = np.fromfunction(lambda i, j: np.power(1 + alpha, np.abs(i + 1) / 12), (br, samples), dtype=float)
    g = np.flipud(g)

    # now join the two arrays
    a = np.vstack((g, h))

    if start_offset > 0:
        a = a[start_offset:]
    if end_offset > 0:
        a = a[:-end_offset]

    return a


class ParameterScenarioSet(object):
    """
    The set of all version of a parameter for all the scenarios.
    """
    default_scenario = 'default'

    "the name of the parameters in this set"
    parameter_name: str
    scenarios: Dict[str, Parameter]

    def __init__(self):
        self.scenarios = {}

    def add_scenario(self, parameter: 'Parameter', scenario_name: str = default_scenario):
        """
        Add a scenario for this parameter.

        :param scenario_name:
        :param parameter:
        :return:
        """
        self.scenarios[scenario_name] = parameter

    def __getitem__(self, item):
        return self.scenarios.__getitem__(item)

    def __setitem__(self, key, value):
        return self.scenarios.__setitem__(key, value)


class ParameterRepository(object):
    """
    Contains all known parameter definitions (so that it is not necessary to re-read the excel file for repeat param accesses).
    The param definitions are independent from the sampling (the Param.__call__ method). Repeat access to __call__ will
    create new samples.

    Internally, parameters are organised together with all the scenario variants in a single ParameterScenarioSet.

    """
    parameter_sets: Dict[str, ParameterScenarioSet]
    tags: Dict[str, Dict[str, Set[Parameter]]]

    def __init__(self):
        self.parameter_sets = defaultdict(ParameterScenarioSet)
        self.tags = defaultdict(lambda: defaultdict(set))

    def add_all(self, parameters: List[Parameter]):
        for p in parameters:
            self.add_parameter(p)

    def clear_cache(self):
        for p_sets in self.parameter_sets.values():
            for param_name, param in p_sets.scenarios.items():
                param.cache = None

    def add_parameter(self, parameter: Parameter):
        """
        A parameter can have several scenarios. They are specified as a comma-separated list in a string.
        :param parameter:
        :return:
        """

        # try reading the scenarios from the function arg or from the parameter attribute
        scenario_string = parameter.source_scenarios_string
        if scenario_string:
            _scenarios = [i.strip() for i in scenario_string.split(',')]
            self.fill_missing_attributes_from_default_parameter(parameter)

        else:
            _scenarios = [ParameterScenarioSet.default_scenario]

        for scenario in _scenarios:
            parameter.scenario = scenario
            self.parameter_sets[parameter.name][scenario] = parameter

        # record all tags for this parameter
        if parameter.tags:
            _tags = [i.strip() for i in parameter.tags.split(',')]
            for tag in _tags:
                self.tags[tag][parameter.name].add(parameter)

    def fill_missing_attributes_from_default_parameter(self, param):
        """
        Empty fields in Parameter definitions in scenarios are populated with default values.

        E.g. in the example below, the source for the Power_TV variable in the 8K scenario would also be EnergyStar.

        +----------+----------+-----+--------+------------+
        | name     | scenario | val | tags   | source     |
        +----------+----------+-----+--------+------------+
        | Power_TV |          | 60  | UD, TV | EnergyStar |
        | Power_TV | 8K       | 85  | new_tag|            |
        +----------+----------+-----+--------+------------+

        **Note** tags must not differ. In the example above, the 8K scenario variable the tags value would be overwritten
        with the default value.

        :param param:
        :return:

        """
        if not self.exists(param.name) or ParameterScenarioSet.default_scenario not in self.parameter_sets[
            param.name].scenarios.keys():
            logger.warning(
                f'No default value for param {param.name} found.')
            return
        default = self.parameter_sets[param.name][ParameterScenarioSet.default_scenario]
        for att_name, att_value in default.__dict__.items():
            if att_name in ['unit', 'label', 'comment', 'source', 'tags']:

                if att_name == 'tags' and default.tags != param.tags:
                    logger.warning(
                        f'For param {param.name} for scenarios {param.source_scenarios_string}, '
                        f'tags is different from default parameter tags. Overwriting with default values.')
                    setattr(param, att_name, att_value)

                if not getattr(param, att_name):
                    logger.debug(
                        f'For param {param.name} for scenarios {param.source_scenarios_string}, '
                        f'populating attribute {att_name} with value {att_value} from default parameter.')

                    setattr(param, att_name, att_value)

    def __getitem__(self, item) -> Parameter:
        """
        Return the default scenario parameter for a given variable name
        :param item: the name of the variable
        :return:
        """
        return self.get_parameter(item, scenario_name=ParameterScenarioSet.default_scenario)

    def get_parameter(self, param_name, scenario_name=ParameterScenarioSet.default_scenario) -> Parameter:
        if self.exists(param_name, scenario=scenario_name):
            return self.parameter_sets[param_name][scenario_name]

        try:
            return self.parameter_sets[param_name][ParameterScenarioSet.default_scenario]
        except KeyError:
            raise KeyError(f"{param_name} not found")

    def find_by_tag(self, tag) -> Dict[str, Set[Parameter]]:
        """
        Get all registered dicts that are registered for a tag

        :param tag: str - single tag

        :return: a dict of {param name: set[Parameter]} that contains all ParameterScenarioSets for all parameter names with a given tag

        """
        return self.tags[tag]

    def exists(self, param, scenario=None) -> bool:
        # if scenario is not None:
        #     return
        present = param in self.parameter_sets.keys()
        if not present:
            return False
        scenario = scenario if scenario else ParameterScenarioSet.default_scenario

        return scenario in self.parameter_sets[param].scenarios.keys()

    def list_scenarios(self, param):
        if param in self.parameter_sets.keys():
            return self.parameter_sets[param].scenarios.keys()


class TableHandler(object):
    version: int

    def __init__(self, version=2):
        self.version = version

    @abstractmethod
    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        raise NotImplementedError()


class Xlsx2CsvHandler(TableHandler):
    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        from xlsx2csv import Xlsx2csv
        data = Xlsx2csv(filename, inmemory=True).convert(None, sheetid=0)

        definitions = []

        _sheet_names = [sheet_name] if sheet_name else [data.keys()]

        for _sheet_name in _sheet_names:
            sheet = data[_sheet_name]

            header = sheet.header
            if header[0] != 'variable':
                continue

            for row in sheet.rows:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell
                definitions.append(values)
        return definitions


class DictReaderStrip(csv.DictReader):
    @property
    def fieldnames(self):
        if self._fieldnames is None:
            # Initialize self._fieldnames
            # Note: DictReader is an old-style class, so can't use super()
            csv.DictReader.fieldnames.fget(self)
            if self._fieldnames is not None:
                self._fieldnames = [name.strip() for name in self._fieldnames]
        return self._fieldnames


class CSVHandler(TableHandler):
    def load_definitions(self, sheet_name, filename=None, id_flag=False):

        reader = DictReaderStrip(open(filename, encoding='utf-8-sig'), delimiter=',')

        definitions = []

        _definition_tracking = defaultdict(dict)

        for i, row in enumerate(reader):

            values = {k: v.strip() for k, v in row.items()}

            if not values['variable']:
                logger.debug(f'ignoring row {i}: {row}')
                continue
            number_columns = []
            if self.version == 2:
                number_columns = ['ref value', 'initial_value_proportional_variation', 'mean growth',
                                  'variability growth']

            if self.version == 1:
                number_columns = ['param 1',
                                  'param 2',
                                  'param 3',
                                  'CAGR']
            for key in number_columns:
                try:
                    if key in values:
                        # guard against empty strings
                        new_val = float(values.get(key, 0) or 0)
                        values[key] = new_val
                except:
                    if 'type' in values and values['type'] == 'interp':
                        # this is a json array ... @todo can we have more validation on these strings?
                        continue
                    else:
                        raise Exception(
                            f'Could not convert value <{values[key]}> for key {key} to number in row {i} for variable {values["variable"]}')

            if 'ref date' in values and values['ref date']:
                if isinstance(values['ref date'], str):
                    values['ref date'] = datetime.datetime.strptime(values['ref date'], '%d/%m/%Y')
                    if values['ref date'].day != 1:
                        logger.warning(
                            f'ref date truncated to first of month for variable {values["variable"]}')
                        values['ref date'] = values['ref date'].replace(day=1)
                else:
                    raise Exception(
                        f"{values['ref date']} for variable {values['variable']} is not a date - "
                        f"check spreadsheet value is a valid day of a month")
            logger.debug(f'values for {values["variable"]}: {values}')
            definitions.append(values)
            scenario = values['scenario'] if values['scenario'] else "n/a"

            if scenario in _definition_tracking[values['variable']]:

                logger.error(
                    f"Duplicate entry for parameter "
                    f"with name <{values['variable']}> and <{scenario}> scenario in file")
                raise ValueError(
                    f"Duplicate entry for parameter "
                    f"with name <{values['variable']}> and <{scenario}> scenario in file")

            else:
                _definition_tracking[values['variable']][scenario] = 1
        return definitions


class PandasCSVHandler(TableHandler):

    def strip(self, text):
        try:
            return text.strip()
        except AttributeError:
            return text

    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        self.version = 2

        import pandas as pd
        df = pd.read_csv(filename, usecols=range(15), index_col=False, parse_dates=['ref date'],
                         dtype={'initial_value_proportional_variation': 'float64'},
                         dayfirst=True,
                         # date_parser=l0ambda x: pd.datetime.strptime(x, '%d-%m-%Y')
                         )
        df = df.dropna(subset=['variable', 'ref value'])
        df.fillna("", inplace=True)

        return df.to_dict(orient='records')


class OpenpyxlTableHandler(TableHandler):
    version: int

    def __init__(self, version=2):
        super().__init__(version=version)
        self.highest_id = -1
        self.id_map = defaultdict(lambda: defaultdict(dict))
        self.id_column = None

    @staticmethod
    def get_sheet_range_bounds(filename, sheet_name):
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows())
        return len(rows)

    def add_ids(self, ws=None, values=None, definitions=None, row_idx=None, id_flag=False,
                sheet_name=None, **kwargs):
        """
        using the id map, assign ids to those variables that have not got an id yet
        :return:
        :rtype:
        """
        name = values["variable"]
        scenario = values['scenario'] if values['scenario'] else "default"
        if name not in self.id_map.keys() or scenario not in self.id_map[name].keys():
            # If this is the first process and it has no ID, set it to 0
            pid = self.highest_id + 1  # else set it to the highest existing ID plus 1
            self.highest_id += 1
            self.id_map[name][scenario] = pid
            values["id"] = pid
            logger.debug(f'{name} {scenario}: {values}')
            definitions[name][scenario]["id"] = pid
            c = ws.cell(row=row_idx + 2, column=self.id_column)
            c.value = pid
            logger.info("ID " + str(pid) + " given to process " + values['variable'])

    def ref_date_handling(self, values: Dict = None, definitions=None, sheet_name=None, id_flag=None,
                          **kwargs):

        if 'ref date' in values and values['ref date']:
            if isinstance(values['ref date'], datetime.datetime):
                # values['ref date'] = datetime.datetime(*xldate_as_tuple(values['ref date'], wb.datemode))
                if values['ref date'].day != 1:
                    logger.warning(f'ref date truncated to first of month for variable {values["variable"]}')
                    values['ref date'] = values['ref date'].replace(day=1)
            else:
                raise Exception(
                    f"{values['ref date']} for variable {values['variable']} is not a date - "
                    f"check spreadsheet value is a valid day of a month")

        logger.debug(f'values for {values["variable"]}: {values}')
        name = values['variable']
        scenario = values['scenario'] if values['scenario'] else "default"
        # store id's in a map to identify largest existing id
        if id_flag:
            if 'id' in values.keys() and (values["id"] or values["id"] == 0):
                pid = values['id']
                if name not in self.id_map.keys() or scenario not in self.id_map[name].keys():
                    # raises exception if the ID already exists
                    if (any(pid in d.values() for d in self.id_map.values())):
                        raise Exception("Duplicate ID variable " + name)
                    else:
                        self.id_map[name][scenario] = pid
                        if pid > self.highest_id:
                            self.highest_id = pid
        if scenario in definitions[name].keys():
            logger.error(
                f"Duplicate entry for parameter "
                f"with name <{values['variable']}> and <{scenario}> scenario in sheet {sheet_name}")
            raise ValueError(
                f"Duplicate entry for parameter "
                f"with name <{values['variable']}> and <{scenario}> scenario in sheet {sheet_name}")
        else:
            definitions[name][scenario] = values

    def table_visitor(self, wb: Workbook = None, sheet_names: List[str] = None, visitor_function: Callable = None,
                      definitions=None, id_flag=False):
        """
        stub for id management

        :param definitions:
        :param wb:
        :type wb:
        :param sheet_names:
        :type sheet_names:
        :param visitor_function:
        :type visitor_function:
        :return:
        :rtype:
        """
        if not sheet_names:
            sheet_names = wb.sheetnames
        for _sheet_name in sheet_names:
            if _sheet_name == 'metadata':
                continue
            sheet = wb[_sheet_name]
            rows = list(sheet.iter_rows())
            header = [cell.value for cell in rows[0]]
            if header[0] != 'variable':
                continue
            if id_flag:
                # get the id column number
                self.id_column = header.index('id') + 1

            for i, row in enumerate(rows[1:]):
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                if not values['variable']:
                    logger.debug(f'ignoring row {i}: {row}')
                    continue

                visitor_function(ws=sheet, values=values, definitions=definitions,
                                 row_idx=i, sheet_name=_sheet_name, id_flag=id_flag, row=row,
                                 header=header)
        return definitions

    def load_definitions(self, sheet_name, filename: str = None, id_flag=False):
        """
        @todo - document that this not only loads definitions, but also writes the data file, if 'id' flag is True
        :param sheet_name:
        :param filename:
        :param id_flag:
        :return:
        """
        import openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)

        definitions = defaultdict(lambda: defaultdict(dict))
        _sheet_names = [sheet_name] if sheet_name else wb.sheetnames
        version = 1

        try:
            sheet = wb['metadata']
            rows = list(sheet.iter_rows())
            for row in rows:
                if row[0].value == 'version':
                    version = row[1].value
            self.version = version
        except:
            logger.info(f'could not find a sheet with name "metadata" in workbook. defaulting to v2')

        table_visitor_partial = partial(self.table_visitor, wb=wb, sheet_names=_sheet_names,
                                        definitions=definitions, id_flag=id_flag)

        table_visitor_partial(visitor_function=self.ref_date_handling)
        if id_flag:
            table_visitor_partial(visitor_function=self.add_ids)
            wb.save(filename)

        res = []
        for var_set in definitions.values():
            for scenario_var in var_set.values():
                res.append(scenario_var)

        return res
        # return [definitions_ .values()]
        # return definitions


class XLWingsTableHandler(TableHandler):
    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        import xlwings as xw
        definitions = []
        wb = xw.Book(fullname=filename)
        _sheet_names = [sheet_name] if sheet_name else wb.sheets
        for _sheet_name in _sheet_names:
            sheet = wb.sheets[_sheet_name]
            range = sheet.range('A1').expand()
            rows = range.rows
            header = [cell.value for cell in rows[0]]

            # check if this sheet contains parameters or if it documentation
            if header[0] != 'variable':
                continue

            total_rows = OpenpyxlTableHandler.get_sheet_range_bounds(filename, _sheet_name)
            range = sheet.range((1, 1), (total_rows, len(header)))
            rows = range.rows
            for row in rows[1:]:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                definitions.append(values)
        return definitions


class TableParameterLoader(object):
    definition_version: int
    """Utility to populate ParameterRepository from spreadsheets.

        The structure of the spreadsheets is:

        | variable | ... |
        |----------|-----|
        |   ...    | ... |

        If the first row in a spreadsheet does not contain they keyword 'variable' the sheet is ignored.

       """

    def __init__(self, filename, table_handler='openpyxl', version=2, **kwargs):
        self.filename = filename
        self.definition_version = version  # default - will be overwritten by handler

        logger.info(f'Using {table_handler} excel handler')
        table_handler_instance = None
        if table_handler == 'csv':
            table_handler_instance = CSVHandler(version)
        if table_handler == 'pandas':
            table_handler_instance = PandasCSVHandler(version)
        if table_handler == 'openpyxl':
            table_handler_instance = OpenpyxlTableHandler()
        if table_handler == 'xlsx2csv':
            table_handler_instance = Xlsx2CsvHandler()
        if table_handler == 'xlwings':
            table_handler_instance = XLWingsTableHandler()
        self.table_handler: TableHandler = table_handler_instance

    def load_parameter_definitions(self, sheet_name: str = None, id_flag=False):
        """
        Load variable text from rows in excel file.
        If no spreadsheet arg is given, all spreadsheets are loaded.
        The first cell in the first row in a spreadsheet must contain the keyword 'variable' or the sheet is ignored.

        Any cells used as titles (with no associated value) are also added to the returned dictionary. However, the
        values associated with each header will be None. For example, given the speadsheet:

        | variable | A | B |
        |----------|---|---|
        | Title    |   |   |
        | Entry    | 1 | 2 |

        The following list of definitions would be returned:

        [ { variable: 'Title', A: None, B: None }
        , { variable: 'Entry', A: 1   , B: 2    }
        ]

        :param sheet_name:
        :return: list of dicts with {header col name : cell value} pairs
        """
        definitions = self.table_handler.load_definitions(sheet_name, filename=self.filename, id_flag=id_flag)
        self.definition_version = self.table_handler.version
        return definitions

    def load_into_repo(self, repository: ParameterRepository = None, sheet_name: str = None, id_flag=False):
        """
        Create a Repo from an excel file.
        :param repository: the repository to load into
        :param sheet_name:
        :return:
        """
        repository.add_all(self.load_parameters(sheet_name, id_flag=id_flag))

    def load_parameters(self, sheet_name, id_flag=False):

        parameter_definitions = self.load_parameter_definitions(sheet_name=sheet_name, id_flag=id_flag)
        params = []

        param_name_map = param_name_maps[int(self.definition_version)]

        for _def in parameter_definitions:

            # substitute names from the headers with the kwargs names in the Parameter and Distributions classes
            # e.g. 'variable' -> 'name', 'module' -> 'module_name', etc
            parameter_kwargs_def = {}
            for k, v in _def.items():
                if k in param_name_map:
                    if param_name_map[k]:
                        parameter_kwargs_def[param_name_map[k]] = v
                    else:
                        parameter_kwargs_def[k] = v

            name_ = parameter_kwargs_def['name']
            del parameter_kwargs_def['name']
            p = Parameter(name_, version=self.definition_version, **parameter_kwargs_def)
            params.append(p)
        return params
