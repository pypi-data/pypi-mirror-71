#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from operator import itemgetter

import libimagequant as liq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from tqdm import tqdm

# LIMIT_OF_EXCEL = 0x18001
reduced_ratio = 1.0
LIMIT_OF_EXCEL = int(0x18001 / reduced_ratio)


def ctuple2cstr(tup, dic=None):
    if dic is not None:
        tup = dic[tup]

    r, g, b = tup

    # silly clamp
    if r > 255:
        r = 255
    if g > 255:
        g = 255
    if b > 255:
        b = 255

    return "%02X%02X%02X" % (r, g, b)


def quantize(input_image=None):
    img = input_image.convert('RGBA')
    width = img.width
    height = img.height
    input_rgba_pixels = img.tobytes()

    # Use libimagequant to make a palette for the RGBA pixels

    attr = liq.Attr()
    input_image = attr.create_rgba(input_rgba_pixels, width, height, 0)
    attr.max_colors = 65535
    result = input_image.quantize(attr)

    # Use libimagequant to make new image pixels from the palette

    # result.dithering_level = 1.0

    raw_8bit_pixels = result.remap_image(input_image)
    palette = result.get_palette()

    # Save converted pixels as a PNG file
    # This uses the Pillow library for PNG writing (not part of libimagequant)
    img = Image.frombytes('P', (width, height), raw_8bit_pixels)

    palette_data = []
    for color in palette:
        palette_data.append(color.r)
        palette_data.append(color.g)
        palette_data.append(color.b)
    img.putpalette(palette_data)

    return img


def get_image_color_dict(im, size_im):
    width, height = im.size
    c_width, c_height = size_im
    im2 = im.resize(size_im, resample=Image.LANCZOS)
    if c_width != width or c_height != height:
        print("resized to: size=%s" % str(im2.size))
    bands = im2.getbands()
    cdict = {}
    for ix, b in enumerate(bands):
        cdict[b] = ix

    if ("R" in cdict) and ("G" in cdict) and ("B" in cdict):
        rgb = list(zip(list(im2.getdata(band=cdict["R"])), list(
            im2.getdata(band=cdict["G"])), list(im2.getdata(band=cdict["B"]))))
    else:
        im3 = im2.convert("RGB")
        bands = im3.getbands()
        cdict = {}
        for ix, b in enumerate(bands):
            cdict[b] = ix
        rgb = list(zip(list(im3.getdata(band=cdict["R"])), list(
            im3.getdata(band=cdict["G"])), list(im3.getdata(band=cdict["B"]))))

    rgbmap = {(r, g, b): (r, g, b) for r, g, b in rgb}
    num_c = len(list(rgbmap.keys()))
    print("read colors: %d" % num_c)
    print("reduced colors: %d" % int(num_c * reduced_ratio))

    return rgb, rgbmap, num_c


def resize_tuple(width, height):
    all_pixel = width * height
    if all_pixel < LIMIT_OF_EXCEL:
        return (width, height)
    alpha = (LIMIT_OF_EXCEL / all_pixel) ** 0.5
    print("alpha = %f" % alpha)
    r_width, r_height = int(width * alpha), int(height * alpha)
    actual_pixel = r_width * r_height
    print("actual pixel to write = %d" % actual_pixel)

    return (r_width, r_height)


def get_reduced_color_map(num_c, rgbmap):
    if reduced_ratio == 1.0:
        return rgbmap

    spl = int((num_c * reduced_ratio) ** 0.3333333333333333)
    spl1p = spl + 1
    st = int(256 / spl)

    index_dict = {}
    for i in range(spl1p):
        for j in range(spl1p):
            for k in range(spl1p):
                index_dict[i * spl1p * spl1p + j * spl1p + k] = 0

    new_color_dict = {}
    for idx, s in enumerate(list(rgbmap.keys())):
        r, g, b = s
        mr, mg, mb = int(
            ((r + 0.5) / 256) * 256), int(((g + 0.5) / 256) * 256), int(((b + 0.5) / 256) * 256)
        i, j, k = int(r / 256 * st), int(g /
                                         256 * st), int(b / 256 * st)
        index = i * spl1p * spl1p + j * spl1p + k
        if index_dict[index] == 0:
            index_dict[index] = (
                int((r / st + 0.5) * st), int((g / st + 0.5) * st), int((b / st + 0.5) * st))
        new_color_dict[s] = index_dict[index]

    return new_color_dict


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    height_in_points = 10.0
    width_in_charwidth = 1.6

    wss = []
    for fidx, f in enumerate(inputs):
        if fidx == 0:
            ws = wb.active
            ws.title = os.path.basename(f)
        else:
            ws = wb.create_sheet(os.path.basename(f))

        with Image.open(f) as im:
            width, height = im.size
            print("%s: format=%s, size=%s" % (f, im.format, im.size))
            resized = im.size
            rgb, rgbmap, num_c = get_image_color_dict(im, resized)

            if num_c > 65535:
                # attempt to libimagequant
                im2 = quantize(input_image=im)
                rgb, rgbmap, num_c = get_image_color_dict(im2, resized)

                # if num_c > 65535:
                #     resized = resize_tuple(width, height)
                #     rgb, rgbmap, num_c = get_image_color_dict(im, resized)

                #     stepx = int(width / 100.0)
                #     stepy = int(height / 100.0)

                #     if num_c > 65535:
                #         while num_c > 65535:
                #             resized = (resized[0] - stepx, resized[1] - stepy)
                #             rgb, rgbmap, num_c = get_image_color_dict(
                #                 im, resized)
                #     else:
                #         while num_c <= 65535:
                #             oldtup = rgb, rgbmap, num_c, resized
                #             resized = (resized[0] + stepx, resized[1] + stepy)
                #             rgb, rgbmap, num_c = get_image_color_dict(
                #                 im, resized)

                #         rgb, rgbmap, num_c, resized = oldtup

            c_width, c_height = resized

            new_color_dict = get_reduced_color_map(num_c, rgbmap)
            fo = Font(name='Calibri', size=1, color="FFFFFFFF")
            bytes_written = 0
            for x in tqdm(range(c_width)):
                for y in range(c_height):
                    color = ctuple2cstr(
                        rgb[y * c_width + x], dic=new_color_dict)
                    c = ws.cell(column=(x + 1), row=(y + 1),
                                value=("#%s" % color))
                    bytes_written += 1
                    co = Color(color)
                    c.font = fo
                    c.fill = PatternFill(fgColor=co, fill_type="solid")

            for y in range(c_height):
                ws.row_dimensions[y + 1].height = height_in_points

            for x in range(c_width):
                ws.column_dimensions[get_column_letter(
                    x + 1)].width = width_in_charwidth

    wb.save(output)
