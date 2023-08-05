#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from operator import itemgetter
from time import time

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from sklearn.cluster import KMeans
from sklearn.datasets import load_sample_image
from sklearn.metrics import pairwise_distances_argmin
from sklearn.utils import shuffle
from tqdm import tqdm

# LIMIT_OF_EXCEL = 0x18001
reduced_ratio = 1.0
LIMIT_OF_EXCEL = int(0x18001 / reduced_ratio)


def ctuple2cstr(tup, dic=None):
    if dic is not None:
        tup = dic[tup]

    r, g, b = tup

    # silly clamp
    if r > 255:
        r = 255
    if g > 255:
        g = 255
    if b > 255:
        b = 255

    return "%02X%02X%02X" % (r, g, b)


def quantize(rgb=None, rgbmap=None):
    from sklearn.cluster import KMeans
    import numpy as np
    colorlist = list(rgbmap.keys())
    colors = len(colorlist)
    X = np.array(colorlist)
    clusters = 65535
    kmeans = KMeans(n_clusters=clusters, verbose=2,
                    random_state=0, init='k-means++', algorithm='full', n_init=1).fit(X)
    result = kmeans.labels_
    centers = kmeans.cluster_centers_
    clist = [[] for x in range(colors)]
    for idx, r in enumerate(result):
        clist[r].append(idx)

    remapped = {}
    for idx, tup in enumerate(colorlist):
        r, g, b = centers[idx]
        remapped[tup] = (r, g, b)

    new_rgb = []
    for tup in rgb:
        new_rgb.append(remapped[tup])

    return new_rgb, remapped, clusters

# Authors: Robert Layton <robertlayton@gmail.com>
#          Olivier Grisel <olivier.grisel@ensta.org>
#          Mathieu Blondel <mathieu@mblondel.org>
#
# License: BSD 3 clause

# tweaked by Takeshi Kimura


def recreate_image(rgb=None, w=0, h=0, n_colors=32767):
    """Recreate the (compressed) image from the code book & labels"""
    # Load the Summer Palace photo

    # china = load_sample_image(filename)

    # Convert to floats instead of the default 8 bits integer coding. Dividing by
    # 255 is important so that plt.imshow behaves works well on float data (need to
    # be in the range [0-1])
    china = np.array(rgb, dtype=np.float64) / 255

    image_array = china

    print("Fitting model on a small sub-sample of the data")
    t0 = time()
    image_array_sample = shuffle(image_array, random_state=0)[:n_colors * 2]
    kmeans = KMeans(n_clusters=n_colors, verbose=2, random_state=0, init='k-means++',
                    algorithm='full', n_init=1, max_iter=30, n_jobs=2).fit(image_array_sample)
    print("done in %0.3fs." % (time() - t0))
    # Get labels for all points
    print("Predicting color indices on the full image (k-means)")
    t0 = time()
    labels = kmeans.predict(image_array)
    print("done in %0.3fs." % (time() - t0))

    codebook = kmeans.cluster_centers_

    d = codebook.shape[1]
    rgb = []
    label_idx = 0
    for j in range(h):
        for i in range(w):
            r, g, b = codebook[labels[label_idx]].tolist()
            rgb.append((int(r * 255), int(g * 255), int(b * 255)))
            label_idx += 1

    return rgb, len(codebook)


def get_image_color_dict(im, size_im):
    width, height = im.size
    c_width, c_height = size_im
    im2 = None
    if (c_width == width) and (c_height == height):
        im2 = im
    else:
        im2 = im.resize(size_im, resample=Image.LANCZOS)
        print("resized to: size=%s" % str(im2.size))
    bands = im2.getbands()
    cdict = {}
    for ix, b in enumerate(bands):
        cdict[b] = ix

    if ("R" in cdict) and ("G" in cdict) and ("B" in cdict):
        rgb = list(zip(list(im2.getdata(band=cdict["R"])), list(
            im2.getdata(band=cdict["G"])), list(im2.getdata(band=cdict["B"]))))
    else:
        im3 = im2.convert("RGB")
        bands = im3.getbands()
        cdict = {}
        for ix, b in enumerate(bands):
            cdict[b] = ix
        rgb = list(zip(list(im3.getdata(band=cdict["R"])), list(
            im3.getdata(band=cdict["G"])), list(im3.getdata(band=cdict["B"]))))

    rgbmap = {(r, g, b): (r, g, b) for r, g, b in rgb}
    num_c = len(list(rgbmap.keys()))
    print("read colors: %d" % num_c)
    print("reduced colors: %d" % int(num_c * reduced_ratio))

    return rgb, rgbmap, num_c


def resize_tuple(width, height):
    all_pixel = width * height
    if all_pixel < LIMIT_OF_EXCEL:
        return (width, height)
    alpha = (LIMIT_OF_EXCEL / all_pixel) ** 0.5
    print("alpha = %f" % alpha)
    r_width, r_height = int(width * alpha), int(height * alpha)
    actual_pixel = r_width * r_height
    print("actual pixel to write = %d" % actual_pixel)

    return (r_width, r_height)


def get_reduced_color_map(num_c, rgbmap):
    if reduced_ratio == 1.0:
        return rgbmap

    spl = int((num_c * reduced_ratio) ** 0.3333333333333333)
    spl1p = spl + 1
    st = int(256 / spl)

    index_dict = {}
    for i in range(spl1p):
        for j in range(spl1p):
            for k in range(spl1p):
                index_dict[i * spl1p * spl1p + j * spl1p + k] = 0

    new_color_dict = {}
    for idx, s in enumerate(list(rgbmap.keys())):
        r, g, b = s
        mr, mg, mb = int(
            ((r + 0.5) / 256) * 256), int(((g + 0.5) / 256) * 256), int(((b + 0.5) / 256) * 256)
        i, j, k = int(r / 256 * st), int(g /
                                         256 * st), int(b / 256 * st)
        index = i * spl1p * spl1p + j * spl1p + k
        if index_dict[index] == 0:
            index_dict[index] = (
                int((r / st + 0.5) * st), int((g / st + 0.5) * st), int((b / st + 0.5) * st))
        new_color_dict[s] = index_dict[index]

    return new_color_dict


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    height_in_points = 10.0
    width_in_charwidth = 1.6

    wss = []
    for fidx, f in enumerate(inputs):
        if fidx == 0:
            ws = wb.active
            ws.title = os.path.basename(f)
        else:
            ws = wb.create_sheet(os.path.basename(f))

        with Image.open(f) as im:
            width, height = im.size
            print("%s: format=%s, size=%s" % (f, im.format, im.size))
            resized = im.size
            rgb, rgbmap, num_c = get_image_color_dict(im, resized)

            if num_c > 65535:
                rgb, num_c = recreate_image(rgb, w=resized[0], h=resized[1])
                rgbmap = None

            fo = Font(name='Calibri', size=1, color="FFFFFFFF")
            bytes_written = 0

            for x in tqdm(range(width)):
                for y in range(height):
                    color = ctuple2cstr(
                        rgb[y * width + x], dic=rgbmap)
                    c = ws.cell(column=(x + 1), row=(y + 1),
                                value=("#%s" % color))
                    bytes_written += 1
                    co = Color(color)
                    c.font = fo
                    c.fill = PatternFill(fgColor=co, fill_type="solid")

            for y in range(height):
                ws.row_dimensions[y + 1].height = height_in_points

            for x in range(width):
                ws.column_dimensions[get_column_letter(
                    x + 1)].width = width_in_charwidth

    wb.save(output)
