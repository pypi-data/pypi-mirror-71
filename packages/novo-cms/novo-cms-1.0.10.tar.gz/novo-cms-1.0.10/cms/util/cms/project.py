# -*- encoding: utf8 -*-
"""
    CMS SubProject Crawler
"""
import os
import re
import sys
import time
import json
import math
import logging
import textwrap
import codecs

from collections import defaultdict

import openpyxl

import requests

from .title import title, stage_title
from . import parse_config, parse_report_data, write_title

from cms.util import get_logger, User

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__))))

DEFAULT_CONFIGS = [
    './cms.config.xlsx',
    os.path.join(os.path.expanduser('~'), 'cms.config.xlsx'),
    os.path.join(ROOT_DIR, 'conf', 'cms.config.xlsx'),
]

for each in DEFAULT_CONFIGS:
    if os.path.isfile(each):
        DEFAULT_CONFIG = each
        break


reload(sys)
sys.setdefaultencoding('utf-8')


class CMS(object):
    base_url = 'http://cms.novogene.com'

    def __init__(self, username, password, logger=None):
        self.username = username
        self.password = password
        self.logger = logger or get_logger()
        self.session = requests.session()

    def login(self):
        url = self.base_url + '/core/login/login!login.action'
        formdata = {
            'loginInfo.usercode': self.username,
            'loginInfo.userpass': self.password,
        }
        resp = self.session.post(url, data=formdata)
        loginMessage = resp.json()['loginMessage']
        if loginMessage == 'success':
            self.logger.info('Login successfully!')
            return True
        self.logger.error('Login failed as: {}'.format(loginMessage))
        return False

    def selectProjectInfosByCond(self, username, projectnum=None):
        """
            项目列表
        """
        url = self.base_url + '/nhzy/project/project!selectProjectInfosByCond.action'

        formdata = {
            'cond.subprojectoperatorcode': username,
            'cond.kftype': 'A',
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        if projectnum:
            formdata['jsonString'] = '{"cond":{"projectnum":"%s"}}' % projectnum

        resp = self.session.post(url, data=formdata).json()
        return resp['projectInfos']

    def selectSubprojectInfosByCond(self, projectnum=None, username=None):
        """
            子项目列表
            projectnum=None  返回所有子项目列表
            projectnum!=None 返回指定子项目信息
        """
        url = self.base_url + '/nhzy/subproject/subproject!selectSubprojectInfosByCond.action'
        formdata = {
            'cond.batchsubprojectnum': 1,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }

        if projectnum:
            formdata['cond.projectnum'] = projectnum

        if username:
            formdata['cond.subprojectoperatorcode'] = username

        resp = self.session.post(url, data=formdata).json()
        return resp['subprojectInfos']

    def selectSubprojectInfoById(self, subprojectid):
        """
            子项目列表
        """
        url = self.base_url + '/nhzy/subproject/subproject!selectSubprojectInfoById.action'
        formdata = {
            "jsonString": "{subprojectInfo:{subprojectid:'%s'}}" % subprojectid
        }
        resp = self.session.post(url, data=formdata).json()
        return resp['subprojectInfo'][0]

    def selectStageInfosByCond(self, subprojectnum):
        """
            分期查询
        """
        url = self.base_url + '/nhzy/settlementbill/stage!selectStageInfosByCond.action'
        formdata = {
            'cond.subprojectcode': subprojectnum,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.post(url, data=formdata).json()

        return resp['stageInfos']

    def selectContractInfoByContractsno(self, contractno):
        """
            合同查询
        """
        url = self.base_url + '/crm/contract/contract!selectContractInfoByContractsno.action'
        formdata = {
            'cond.contractsno_equals': contractno,
        }
        resp = self.session.post(url, data=formdata).json()

        if len(resp['contractInfo']) != 1:
            self.logger.warn('not unique contract number: {}'.format(contractno))
            exit(1)

        return resp['contractInfo'][0]

    def selectQuotationproductInfosByCond(self, quotationid):
        """
            合同报价查询
        """
        url = self.base_url + '/nhzy/projectquotation/quotationproduct!selectQuotationproductInfosByCond.action'
        formdata = {
            'cond.kfquotationid': quotationid,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quotationproductInfos']

    def selectQuoprocessInfosByCond(self, kfquotationproductid):
        """
            工序查询
        """
        url = self.base_url + '/nhzy/projectquotation/quoprocess!selectQuoprocessInfosByCond.action'
        formdata = {
            'cond.kfquotationproductid': kfquotationproductid,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quoprocessInfos']

    def get_quotation_data(self, contractno):
        quotation_data = {}

        contract = self.selectContractInfoByContractsno(contractno)
        quotation_products = self.selectQuotationproductInfosByCond(contract['quotationid'])

        for quotation in quotation_products:
            pcode = quotation['pcode']

            if pcode not in quotation_data:
                quotation_data[pcode] = defaultdict(list)

                # 上机数据量，样本数
                quotation_data[pcode]['samplenum'] = 0
                quotation_data[pcode]['totalproductdata'] = 0

            # 工序列表
            quoprocesses = self.selectQuoprocessInfosByCond(quotation['kfquotationproductid'])
            for quoprocess in quoprocesses:

                if '上机' in quoprocess['processcode']:
                    quotation_data[pcode]['totalproductdata'] += float(
                        quoprocess['datasize']) * int(quoprocess['samnum'])

                    quotation_data[pcode]['samplenum'] += int(quoprocess['samnum'])

                if quoprocess['datasize'] and quoprocess['datasize'].isdigit():
                    quoprocess['datasize'] = int(quoprocess['datasize'])
                else:
                    quoprocess['datasize'] = ''

                quotation_data[pcode]['processes'] += [
                    {
                        'processcode': quoprocess['processcode'],
                        'samnum': int(quoprocess['samnum']),
                        'datasize': quoprocess['datasize'],
                        'processtypename': quoprocess['processtypename']
                    }
                ]

            quotation_data[pcode]['processtypename'] += quotation['processtypename'].split(',')

        return quotation_data


def main(**args):
    start_time = time.time()

    logger = get_logger('CMS', verbose=args['verbose'])

    msg = []
    for k, v in args.iteritems():
        if k not in ('func', 'subparser_name'):
            if isinstance(v, list):
                v = ', '.join(v)
            msg += ['{:>10}:\t{}'.format(k, v)]
    logger.info('input arguments:\n\033[33m{}\033[0m'.format('\n'.join(msg)))

    # 报表文件解析
    if args['report']:
        logger.info('\033[32mparsing report data: {} ...\033[0m'.format(', '.join(args['report'])))
    report_data = parse_report_data(args['report'])

    # 配置文件解析
    config = parse_config(args['config'])
    if config:
        logger.debug('\033[32mConfig Data: \n{}\033[0m'.format(
            json.dumps(config, ensure_ascii=False, indent=2)))

    # 登录用户名密码
    user = User(**args)
    username, password = user.get_user_pass()
    cms = CMS(username, password, logger=logger)
    if cms.login():
        user.save(username, password)
    else:
        exit(1)

    # 表头设置
    titles = [each[0] for each in title]
    for i in range(1, 6):
        stage_title[0] = '任务{} 任务名称'.format(i)
        titles += stage_title

    # 表头key
    fields = [each[1] for each in title]

    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = unicode('子项目')
    sheet2 = book.create_sheet(unicode('SOP明细'))

    # 写入表头
    write_title(sheet1, sheet2, titles)

    index1 = 0  # sheet1
    index2 = 1  # sheet2

    # 运营项列表
    projects = cms.selectProjectInfosByCond(username, projectnum=args['projectnum'])
    project_count = len(projects)
    logger.info('Total project number of {}: {}'.format(username, project_count))

    # 记录配置文件中没有的工序和产品名称
    sop_not_in_config = set()
    product_not_in_config = set()

    # 子项目列表（可能存在不属于自己部门的子项目...）
    # sub_project_dict = {each['subprojectnum']: 1 for each in cms.selectSubprojectInfosByCond(username=username)}
    # sub_project_count = len(sub_project_dict)
    # logger.info('Total SubProject Number of {}: {}'.format(username, sub_project_count))

    for n, info in enumerate(projects, 1):
        logger.debug('>>> Projct: {projectnum}'.format(**info))

        percent = n * 100. / project_count
        sys.stderr.write(
            '\033[K[{n}/{project_count}] \033[36m{percent:.1f}% \033[0mcompleted\r'.format(**locals()))
        sys.stderr.flush

        # 获取自定项目编号的子项目列表
        sub_projects = cms.selectSubprojectInfosByCond(info['projectnum'])

        # 项目 - 合同 - 工序
        quotation_data = cms.get_quotation_data(info['contractno'])

        for sub_info in sub_projects:  # 每个子项目编号为一行

            # 通过利润中心过滤不属于自己部门的子项目
            if args.get('profitname') and sub_info['profitname'] not in args['profitname']:
                continue

            index1 += 1

            sub_info['index1'] = index1
            logger.debug('> Sub Project: {index1} {subprojectnum}'.format(**sub_info))

            sub_info['contractno'] = info['contractno']

            # 更新：config
            sub_info.update(config['common'])

            if args['PM']:
                sub_info['PM'] = args['PM']

            # 项目税前收入 => 剩余额度
            sub_info_details = cms.selectSubprojectInfoById(sub_info['subprojectid'])
            subprojecttotallimit = float(sub_info_details['subprojecttotallimit'])  # 子项目总额度
            settlementmoney = float(sub_info_details['settlementmoney'])  # 已结算金额
            subprojectleftlimit = subprojecttotallimit - settlementmoney
            sub_info['subprojectleftlimit'] = subprojectleftlimit

            # 更新：样本数，总数据量
            sub_info.update(quotation_data.get(sub_info['pcode'], {}))

            # 使用报表数据进行填充
            if report_data.get(sub_info['subprojectnum']):  # 业务线维度报表
                sub_info['js_samplenum'] = report_data[sub_info['subprojectnum']]['js_samplenum']
                sub_info['js_money'] = report_data[sub_info['subprojectnum']]['js_money']

            if report_data.get((info['projectnum'], sub_info['pcode'])):  # 合同维度结算报表
                result = report_data[(info['projectnum'], sub_info['pcode'])]

                if sub_info.get('totalproductdata') is None:
                    sub_info['totalproductdata'] = result['datasize']

                if sub_info.get('samplenum') is None:
                    sub_info['samplenum'] = result['samplenum']

                if sub_info.get('pname') is None:
                    sub_info['pname'] = result['processtypename']

            # 数据量，样本数 * (剩余额度 / 子项目总额度)
            if not subprojecttotallimit:
                logger.debug('{subprojectnum}的总额度为0'.format(**sub_info))
                limit_percent = 0
            else:
                limit_percent = float(subprojectleftlimit) / subprojecttotallimit

            if sub_info.get('samplenum'):
                sub_info['samplenum'] = int(math.ceil(int(sub_info['samplenum']) * limit_percent))

            if sub_info.get('totalproductdata'):
                sub_info['totalproductdata'] = int(math.ceil(int(sub_info['totalproductdata']) * limit_percent))

            result = config['product'].get(unicode('{pname}'.format(**sub_info)))
            if not result:
                # logger.warn('[{subprojectnum}] {pname} not in config, please add it!'.format(**sub_info))
                product_not_in_config.add((sub_info['subprojectnum'], sub_info['pname']))
                ave_sam_product, pooling, sop_price = 1, 1, 1
            else:
                ave_sam_product, pooling, sop_price = result

            if sub_info['samplenum']:
                sub_info['pc_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(index1+1,
                                                                       ave_sam_product, sub_info['pc_ave_time'])
                sub_info['bi_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(index1+1,
                                                                       ave_sam_product, sub_info['bi_ave_time'])
                sub_info['pc_bi_time'] = '=W{0}+X{0}'.format(index1+1)

            # SOP明细 -- 生产成本，不考虑分析
            # 测序前SOP成本 = SUM(样本数 * 对应单价)
            # 生产SOP成本 = 测序前SOP成本 + 数据量 * 单价 * pooling系数

            if sub_info.get('processes'):
                sopcost_before = []
                sopcost_production = []
                sop_filter = {}
                for process in sub_info['processes']:
                    sop_key = '{processcode}__{processtypename}'.format(**process)

                    # 工序去重
                    if process['processcode'] in sop_filter:
                        continue
                    sop_filter[process['processcode']] = 1
                    
                    index2 += 1

                    ave_sample = config['sop'].get(unicode(sop_key))

                    if ave_sample:
                        logger.debug('{sop_key}: {ave_sample}'.format(**locals()))

                        if '上机' in process['processcode']:  # 测序成本
                            sopcost_production += ['K{}*{}*{}'.format(
                                index1+1, ave_sample, pooling)]
                        elif any([each in process['processcode'] for each in ('提取', '库检', '检测', '建库')]):  # 测序前生产成本
                            sopcost_before += ['J{}*{}'.format(index1+1, ave_sample)]
                    elif '分析' not in process['processcode']:
                        sop_not_in_config.add((sub_info['subprojectnum'], process['processcode'], process['processtypename']))

                    sheet2.cell(row=index2, column=1, value=sub_info['contractno'])
                    sheet2.cell(row=index2, column=2, value=sub_info['subprojectnum'])
                    sheet2.cell(row=index2, column=3, value=process['processcode'])
                    sheet2.cell(row=index2, column=4, value=process['processtypename'])
                    sheet2.cell(row=index2, column=5, value=process['samnum'])
                    sheet2.cell(row=index2, column=6, value=process['datasize'])

                if sopcost_before:
                    sub_info['sopcost_before'] = '=' + ' + '.join(sopcost_before)
                else:
                    sub_info['sopcost_before'] = 0

                if sopcost_production:
                    sub_info['sopcost_production'] = '=L{} + '.format(
                        index1+1) + ' + '.join(sopcost_production)
                else:
                    sub_info['sopcost_production'] = 0
            else:  # 对于没有sop的项目, 直接用单价*样本数 ？？？
                sub_info['sopcost_before'] = '=J{i}*{price}'.format(i=index1+1, price=sop_price)
                sub_info['sopcost_production'] = '=J{i}*{price}'.format(i=index1+1, price=sop_price)


            sub_info['compute_cost'] = '=K{i}*{gcluster}'.format(i=index1+1, **sub_info)
            sub_info['pc_bi_cost'] = '=W{i}*{pccost} + X{i}*{bicost}'.format(i=index1+1, **sub_info)

            sub_info['people_cost'] = '=I{i}/1.06*{apportion_people}'.format(i=index1+1, **sub_info)
            sub_info['three_cost'] = '=I{i}/1.06*{apportion_three}'.format(i=index1+1, **sub_info)

            sub_info['profit'] = '=I{i}/1.06 - M{i} - N{i} - P{i} -Q{i} - R{i}'.format(i=index1+1)

            if subprojectleftlimit:
                sub_info['profit_rate'] = '=S{i}/I{i}'.format(i=index1+1)

            linelist = [sub_info.get(k, '') for k in fields]

            stages = cms.selectStageInfosByCond(sub_info['subprojectnum'])

            for stage_info in stages:
                tmp_list = [''] * 11
                tmp_list[0] = stage_info['installmentcode']
                tmp_list[1] = stage_info['operatemanagerdesc']
                tmp_list[3] = stage_info['informationleader']
                tmp_list[6] = stage_info['desc12']
                linelist += tmp_list

            for col, value in enumerate(linelist, 1):
                try:
                    value = float(value)
                except:
                    try:
                        value = int(value)
                    except:
                        pass

                if isinstance(value, list):
                    value = ','.join(value)
                sheet1.cell(row=index1 + 1, column=col, value=value)

        if args['limit'] and index1 >= args['limit']:
            break

    sys.stderr.write('\n\n')

    if sop_not_in_config:
        logger.warn('there are some sops not in config, please check:')
        with codecs.open('sop_not_in_config.xls', 'w', encoding='gbk') as out:
            for each in sop_not_in_config:
                line = '\t'.join(each)
                print(line)
                out.write(line + '\n')

    if product_not_in_config:
        logger.warn('there are some sops not in config, please check:')
        with codecs.open('product_not_in_config.xls', 'w', encoding='gbk') as out:
            for each in product_not_in_config:
                line = '\t'.join(each)
                print(line)
                out.write(line + '\n')

    book.save(args['out'])
    logger.info('\033[33msave file: {out}\033[0m'.format(**args))
    logger.info('\033[32mtime used: {:.1f}s\033[0m'.format(time.time() - start_time))


def parser_add_cms(parser):

    parser.description = __doc__
    parser.epilog = textwrap.dedent('''
        \033[36mexamples:
            %(prog)s -o disease.xlsx
            %(prog)s -o disease.xlsx -r cms_report.xlsx      # 使用CMS导出的报表文件，用于更新结算样本数
            %(prog)s -o disease.xlsx -l 100                  # 只输出100条记录，可用于测试
            %(prog)s -o disease.xlsx -c your_config.xlsx     # 使用指定的配置文件
            %(prog)s -o disease.xlsx -pm Name                # 指定项目经理名字
            %(prog)s -o disease.xlsx -profit 疾病研究部      # 指定利润中心，用于过滤不属于自己部门的项目
        \033[0m
    ''')

    parser.add_argument(
        '-c', '--config', help='the config excel [default: %(default)s]', default=DEFAULT_CONFIG)
    parser.add_argument(
        '-o', '--out', help='the output filename [default: %(default)s]', default='cms.project.xlsx')
    parser.add_argument(
        '-r', '--report', help='the report excel which contains sample informations', nargs='*')

    parser.add_argument('-l', '--limit', help='the limit count of output, just for test', type=int)
    parser.add_argument('-project', '--projectnum', help='specify a project number')

    parser.add_argument('-pm', '--PM', help='the name of PM')
    parser.add_argument('-profit', '--profitname', help='the profitname to filter', nargs='*')

    parser.set_defaults(func=main)
