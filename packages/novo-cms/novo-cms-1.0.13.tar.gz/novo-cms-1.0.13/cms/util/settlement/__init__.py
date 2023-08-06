# -*- encoding: utf8 -*-
import re
import sys
import json
import datetime
import warnings
from collections import defaultdict
from copy import deepcopy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side


warnings.simplefilter("ignore")


def write_title(sheet1, title):

    colors = '64F336 EBBF1D 00FFF3'.split()
    col = 0
    for part, color in zip(title, colors):
        for field in part:
            value = field[0]
            col += 1
            _ = sheet1.cell(row=1, column=col, value=value)
            _.fill = PatternFill(start_color=color,
                                 end_color=color,
                                 fill_type="solid")
            _.font = Font(bold=True)

            sheet1.column_dimensions[get_column_letter(col)].width = 20


def parse_report_data(report_xlsx_list):
    """
        业务线维度 - 结算报表信息提取
    """
    fields = [
        ['子项目编号', 'subprojectnum'],
        ['子项目名称', 'subprojectname'],
        ['子项目产品类型', 'productname'],
        ['运营经理', 'pc_name'],
        ['信息负责人', 'bi_name'],
        ['部门', 'department'],
        ['系统分期编号', 'stagecode'],
        ['结算样本个数', 'samplenum'],
        ['理论结题时间', 'pc_plan_time'],
        ['实际结题时间', 'pc_complete_time'],
        ['理论结题时间', 'pc_plan_time'],
    ]

    for report_xlsx in report_xlsx_list:
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.get_active_sheet()
        for row in ws.rows:
            if row[0].value == '序号':
                title = [each.value.encode() for each in row]
                continue
            elif not isinstance(row[0].value, float):
                continue

            linelist = [each.value for each in row]

            context = dict(zip(title, linelist))

            result = {k2: context[k1] for k1, k2 in fields}

            if not result['samplenum']:
                continue

            result['pc_name'] = result['pc_name'].split()[-1]
            result['samplenum'] = int(result['samplenum'])

            yield result

