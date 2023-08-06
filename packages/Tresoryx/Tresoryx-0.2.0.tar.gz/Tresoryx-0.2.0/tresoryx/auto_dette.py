#!/usr/bin/env python3
# -*- coding:utf-8 -*-

# Copyright 2018 Guitou des Phoenix
#
# Ce fichier fait partie de Tresoryx, sous license GNU GPL.
# This file is part of Tresoryx, under GNU GPL license.


"""Renvoie les avances non remboursées au club, sous forme de tableau 
ainsi qu'un diagramme."""


import datetime as dt
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
import argparse as ap
try:
    import matplotlib as mpl
    mpl_dispo = True
except ImportError:
    mpl_dispo = False

if mpl_dispo:
    try:
        mpl.use('TkAgg')  # Requires the tkinter module
    except ImportError:
        try:
            mpl.use('Qt5Agg')
        except ImportError:
            mpl.use('Agg')

    import matplotlib.pyplot as plt
    mpl.rcParams['grid.alpha'] = 0.5
    mpl.rcParams['grid.linestyle'] = '--'

#import seaborn as sb
from .communs import ChargeurDonnees, normalise_nom
import logging
loggeur = logging.getLogger(__name__)
logging.basicConfig(level=logging.WARNING)
loggeur.setLevel(logging.INFO)

fichier_remboursements = "Remboursements_2018-2019.xlsx"
fichier_exercice = "../auto_exercice/Exercice.xlsx"
#fichier_noms = "../auto_exercice/noms.csv"
feuille_r = 'General'
ncols = 30
feuille_e = 0
source = 'excel'
ligne_debut = 0


# Crée un style pour griser les cellules déjà remboursées
# Suppose que 'Général' est ordonné de gauche à droite par date...
def colorie_rembourse(montants, fait=0):
    """Grise ce qui a été remboursé + en rouge si montants ne correspondent pas."""
    dejapaye = montants[(fait+1):].cumsum() - montants[fait]
    styles = np.array([''] + ['color: grey' if round(m, 2)<=0 else '' for m in dejapaye])
    # Au moins une des dettes cumulées doit tomber sur zéro si les remboursements
    # faits sont corrects.
    if montants[fait] > 0 and (dejapaye.round(2) != 0).all():
        colonne_dejapaye = dejapaye.searchsorted(0)
        if isinstance(colonne_dejapaye, np.ndarray):
            colonne_dejapaye = colonne_dejapaye[0]
        try:
            styles[:colonne_dejapaye+1] = 'color: red'
        except IndexError:
            styles[0] = 'color: red'
    return styles


def vide_rembourse(montants, fait=0):
    valeurs_centrees = montants[(fait+1):].cumsum() - montants[fait]
    nvx_montants = montants.copy()
    nvx_montants[np.concatenate(([False]*(fait+1),
                                 valeurs_centrees.round(2) <= 0))] = np.NaN
    return nvx_montants


def details_vers_excel(dettes, competitions, excel, feuille='Détails'):
    dettes[['dette', 'Total', 'Remboursé', 'Date', 'Intitulé']
           +competitions
           +['_merge', 'Remboursé_autre', 'Date_autre', 'Intitulé_autre']
          ].rename(columns={
                    'dette': 'Reste à rembourser',
                    'Total': "Total sur l'année",
                    'Remboursé': 'A déjà remboursé',
                    'Date': 'Dates',
                    'Intitulé': 'Intitulés'}
          ).style.apply(colorie_rembourse, axis=1, raw=True,
                        subset=['A déjà remboursé'] + competitions
          ).applymap(lambda v: 'color: #D3D3D3' if round(v,2)==0 else '',
                     subset=competitions
          ).applymap(lambda v: ('color: red' if v=='left_only'
                                else 'color: orange' if v=='right_only'
                                else ''),
                     subset=['_merge']
          ).applymap(lambda v: 'color: #191970', subset=['Reste à rembourser']
          ).bar(align='mid', color='#191970', subset=['Reste à rembourser']
          ).to_excel(excel, sheet_name=feuille) #, float_format='%.2f')

    feuille_details = excel.sheets[feuille]
    
    extra_col = 8 + len(competitions)
    loggeur.debug('%d competitions', len(competitions))
    try:
        feuille_details.set_column("A:D", 20)
        feuille_details.set_column("E:E", 13)
        #feuille_details.set_column(range(4), 20)
        feuille_details.set_column('%s:%s' % (get_column_letter(extra_col),
                                              get_column_letter(extra_col+3)),
                                   22)
    except AttributeError:
        # Openpyxl
        for col in "ABCD":
            feuille_details.column_dimensions[col].width = 20
        feuille_details.column_dimensions["E"].width = 13
        for coln in range(extra_col, extra_col+4):
            feuille_details.column_dimensions[get_column_letter(coln)].width = 22


def graphe(dettes, competitions, classe):
    fig, ax = plt.subplots(figsize=(12, 7))

    bases = -dettes['Remboursé'].values
    x = np.arange(dettes.shape[0])
    ax.bar(x, dettes['Remboursé'].values, width=-0.4, bottom=bases,
           color='gray', align='edge', label='Remboursé')

    carte_couleurs = plt.get_cmap('tab20')
    toutes_barres = []
    for i, compet in enumerate(competitions):
        barres = ax.bar(x, dettes[compet].values, width=0.4, bottom=bases,
                        align='edge', label=compet, 
                        color=carte_couleurs(i % carte_couleurs.N))
        bases += dettes[compet].values
        toutes_barres.append(barres)
    
    ax.set_xticks(x)
    ax.set_xticklabels(dettes.index.values, rotation='45',
                       horizontalalignment='right',
                       verticalalignment='top')
    ax.legend()
    #ax.set_ylabel("Dette = Crédit - Remboursements")
    ax.set_title("Dettes et remboursements au " + dt.datetime.now().strftime('%d/%m/%Y'))

    ylim = ax.get_ylim()
    ax.set_ylim(min(0, ylim[0]), max(0, ylim[1]))
    ax.set_ylabel('Dette')
    ax.grid(True)
    #plt.show(block=True)
    fig.savefig('dettes%s.png' % classe, bbox_inches='tight')
    return ax


def auto_dette(fichier_remboursements, fichier_exercice, #fichier_noms,
         feuille_r='General', debut=None, ncols=30, feuille_e=0,
         classe='755', source='excel', sortie='table', cours_sec=False):
    #TODO: specifier date de début dans ChargeurDonnees
    # Lis le fichier Remboursements
    du_excel = pd.ExcelFile(fichier_remboursements)
    # Lis la 1è ligne pour les catégories (Compétitions VS Achats)
    du_categories = du_excel.parse(feuille_r, index_col=False,
                                   header=None,
                                   skiprows=ligne_debut,
                                   nrows=2,
                                   usecols=tuple(range(ncols)))\
                    .fillna(method='ffill',axis=1).T
    du_categories = du_categories[du_categories[1] != 'Total'].set_index(0)[1]
    du = du_excel.parse(feuille_r, index_col=0, header=0,
                        skiprows=ligne_debut+1,
                        usecols=tuple(range(ncols)))
    du_excel.close()

    colonnes_vides = du.columns.to_series().str.startswith('Unnamed: ')
    du.drop(du.columns[colonnes_vides], axis=1, inplace=True)
    du.drop(['compte', 'Total', 'Extés', np.nan], inplace=True)  # Lignes

    #TODO: MultiIndex avec Competition/Matériel
    competitions = du.columns.to_series().drop(['Surnom', 'Total']).tolist()
    du[competitions + ['Total']] = du[competitions + ['Total']].fillna(0)

    totaux_corrects = (du[competitions].sum(axis=1) - du.Total < 0.0001)
    assert totaux_corrects.all(), \
        "La colonne Total ne correspond pas à la somme des compétitions:\n" + \
        str(du.Total[~totaux_corrects])
    du = du[du.Total.round(2) != 0]#.copy()

    # Lis le fichier Exercice
    #exer_excel  = pd.ExcelFile(fichier_exercice)
    #exercice = exer_excel.parse(feuille_e, skiprows=4,
    #                            usecols=list(range(1, 10)))
    #exer_excel.close()
    #exercice.columns = exercice.columns.str.capitalize()
    #colonnes_attendues = np.array(['Date', 'Intitulé', 'Nom', 'Crédit', 'Débit', 'Solde', 'Mode', 'Classe'])
    #if (exercice.columns[:len(colonnes_attendues)] != colonnes_attendues).any():
    #    loggeur.warning('Les colonnes trouvées ne correspondent pas:\n%s\n',
    #                   exercice.head().to_string())

    #exercice.Date = pd.to_datetime(exercice.Date, errors='coerce',
    #                               dayfirst=True).dt.floor('d')

    donnees = ChargeurDonnees(fichier_exercice, feuille_e, source) #, fichier_noms)
    exercice = donnees.exercice
    if debut:
        date_debut = dt.datetime.strptime(debut, '%Y-%m-%d')
        ligne_debut_rbsts = exercice.Date.searchsorted(date_debut)
        # Changed in Pandas version 0.24: returns a scalar.
        if isinstance(ligne_debut_rbsts, np.ndarray):
            ligne_debut_rbsts = ligne_debut_rbsts[0]
        exercice = exercice.iloc[ligne_debut_rbsts:]

    #exercice['Surnom'] = normalise_nom(exercice.Nom, fichier_noms,
    #                                   de='Complet', vers='Surnom')
    # On va plutôt se baser sur le nom complet...

    classes = {'Compétitions': '755', 'Achats': '707'}  # ~~> config
    #toutes_dettes = [] (à rembourser, détails)
    #for categorie, classe in classes.items():
    #    du[['Surnom'] + du_categories[categorie].tolist() + ['Total']]

    # 755 - Remboursement de frais pour service rendu aux adhérents
    classe_rbst = exercice.Classe.fillna('').str.startswith(classe)
    #TODO:
    # 707 - Ventes matériels (livres, brochures, vente aux membres)
    assert not classe_rbst.isna().any()
    rbst_joueurs = exercice[classe_rbst].groupby('Nom')
    rbsts = rbst_joueurs.agg({'Crédit': 'sum',  # '\n'.join
                              'Intitulé': lambda x: '.\n'.join(x.fillna('')),
                              'Date': lambda x: '\n'.join(
                                            x.dt.strftime('%Y-%m-%d'))
                              })
    autres_rbsts = exercice[~classe_rbst
                            & exercice.Classe.fillna('').str.startswith('7')]\
                           .groupby('Nom')\
                           .agg(
                             {'Crédit': lambda x: ' + '.join(x.astype(str)),
                              'Intitulé': lambda x: '.\n'.join(x.fillna('')),
                              'Date': lambda x: '\n'.join(
                                                     x.dt.strftime('%Y-%m-%d'))
                              })

    dettes = pd.merge(rbsts.join(autres_rbsts, rsuffix='_autre'),  #how='outer'
                      du,
                      how='outer', #on='Nom')
                      left_index=True, right_index=True,
                      validate='one_to_one', indicator=True)
                      #left_index=True, right_on='Nom')
    dettes[competitions + ['Crédit', 'Total']] = dettes[competitions + ['Crédit', 'Total']].fillna(0)

    if (dettes._merge == 'left_only').any():
        loggeur.warning('Rembourseurs sans dettes:\n%s.',
            ',\n'.join(dettes.index.to_series()[dettes._merge=='left_only'].sort_values()))
        #print('Rembourseurs non trouvés:\n%s' % \
        #      (dettes.index.to_series()[dettes._merge=='left_only'],))
    if (dettes._merge == 'right_only').any():
        loggeur.info("Endettés n'ayant rien remboursé:\n%s.",
            ',\n'.join(dettes.index.to_series()[dettes._merge=='right_only'].sort_values()))
        #print("Endettés n'ayant rien remboursé:\n%s" % \
        #      (dettes.index.to_series()[dettes._merge=='right_only'],))
    #dettes.index = dettes.Surnom
    
    # Tri par dette
    dettes.rename(columns={'Crédit': 'Remboursé',
                           'Crédit_autre': 'Remboursé_autre'}, inplace=True)
    dettes['dette'] = dettes['Total'] - dettes['Remboursé']  # >0
    dettes['arrondi_dette'] = dettes.dette.round(2)
    dettes['arrondi_remboursé'] = dettes['Remboursé'].round(2)
    dettes = dettes.sort_values(['arrondi_dette', 'arrondi_remboursé'],
                                ascending=False)\
                   .drop(['arrondi_dette', 'arrondi_remboursé'], axis=1)
    #.dette.plot.bar()

    # Supprime les personnes étant à jour
    joueurs_a_jour = dettes.dette.round(2) <= 0
    loggeur.info('Joueurs à jour:\n%s.', '\n'.join(sorted(dettes[joueurs_a_jour].index)))
    dettes_a_jour = dettes[joueurs_a_jour].sort_index()
    dettes = dettes[~joueurs_a_jour]
    compets_a_jour = dettes[competitions].columns[
                            (
                             dettes[competitions].cumsum(axis=1).round(2).T
                             <= dettes['Remboursé']
                            ).all(axis=1)
                        ].tolist()
    loggeur.info('Passe les compétitions à jour:\n%s', '\n'.join(compets_a_jour))
    if not set(competitions).difference(compets_a_jour):
        loggeur.warning('Toutes les compétitions sont à jour.')

    dettes_actuelles = dettes.drop(compets_a_jour, axis=1)
    dettes_actuelles['Remboursé'] -= dettes[compets_a_jour].sum(axis=1)#.round(2)
    compets_actuelles = dettes_actuelles.columns.intersection(competitions).tolist()

    if mpl_dispo:
        graphe(dettes, competitions, classe)
    else:
        loggeur.warning('Matplotlib non installé. Nécessaire pour le graphique.')

    #elif sortie == 'table':
    #dettes.to_csv('dettes.tsv', sep='\t', decimal=',')
    
    dettes_excel = pd.ExcelWriter('dettes%s.xlsx' % classe)
    dettes[['dette', 'Remboursé'] + competitions]\
            .apply(vide_rembourse, axis=1, raw=True, args=(1,),
                   result_type='broadcast')\
            [['dette'] + compets_actuelles]\
            .sort_index()\
            .rename(columns={'dette': 'Reste à rembourser'})\
            .style.applymap(lambda v: 'color: #191970',
                            subset=['Reste à rembourser'])\
            .applymap(lambda v: 'color: #1A1A1A',
                      subset=compets_actuelles)\
            .to_excel(dettes_excel,
                      sheet_name='À rembourser')#,
                      #float_format='%.2f')
    try:
        dettes_excel.sheets['À rembourser'].set_column("A:B", 20)
        dettes_excel.sheets['À rembourser'].set_column("C:Z", 15)
    except AttributeError:
        # Openpyxl
        try:
            dettes_excel.sheets['À rembourser'].column_dimensions["A"].width = 20
            dettes_excel.sheets['À rembourser'].column_dimensions["B"].width = 20
            for col in "CDEFGHIJKLMNOPQRSTUVWXYZ":
                dettes_excel.sheets['À rembourser'].column_dimensions[col].width = 20
        except AttributeError:
            #print(type(dettes_excel.sheets['À rembourser']))
            dettes_excel.sheets['À rembourser'].col(1).set_width(20)
            dettes_excel.sheets['À rembourser'].col(2).set_width(20)
            for coln in range(3,26):
                dettes_excel.sheets['À rembourser'].col(coln).set_width(15)

    details_vers_excel(dettes, competitions, dettes_excel, feuille='Détails')
    details_vers_excel(dettes_a_jour, competitions, dettes_excel, feuille='Joueurs à jour')
    
    dettes_excel.close()
    #else:
    #    raise ValueError('Option de sortie %r non valide.' % sortie)


def main():
    parseur = ap.ArgumentParser(description=__doc__)
    parseur.add_argument('remboursements')
    parseur.add_argument('exercice')
    parseur.add_argument('-f', '--feuille', default='General')
    #parseur.add_argument('-n', '--noms', dest='fichier_noms')
    parseur.add_argument('-t', '--table', action='store_const', const='table',
                        dest='sortie', default='graphe')
    #parseur.add_argument('-g', '--graphe', action='store_const', const='graphe',
    #                    dest='sortie')
    parseur.add_argument('-d', '--debut',
                        help='Date de début (exercice) format %%Y-%%m-%%d.')
    parseur.add_argument('-c', '--classe', type=str, default='755',
                        help='classe de remboursement à vérifier [%(default)s].')
    parseur.add_argument('-n', '--cours-sec', action='store_true',
                        help="Calcule mais n'enregistre rien.")
    #parseur.add_argument('-o', '--oubli', action='store_true')
    args = parseur.parse_args()
    auto_dette(args.remboursements, args.exercice, feuille_r=args.feuille,
               debut=args.debut, classe=args.classe, sortie=args.sortie)


if __name__ == '__main__':
    main()
